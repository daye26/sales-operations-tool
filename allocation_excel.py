from collections import Counter
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
import sys
import unicodedata
import warnings

from openpyxl import Workbook, load_workbook

import asignaciones_excel as preallocation_engine
from excel_sheet_selection import select_active_then_sheet1
import free_cars_history
from excel_output import append_row, calculate_column_widths, prepare_worksheet, save_workbook_atomically
from port_resolution import resolve_port
import tabular_normalization as tabular
from warning_samples import WarningSamples


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

BASE_EXCEL_DIR = Path.cwd()
OUTPUT_XLSX_PATH = BASE_EXCEL_DIR / "vehicle_allocation_result.xlsx"
PROGRESS_CALLBACK = None
SHORT_DATE_FORMAT = "yyyy-mm-dd"
ALLOCATION_WINDOW_MODES = {"eta_days", "gate_in"}
ALLOCATION_WINDOW_MODE = preallocation_engine.PREALLOCATION_WINDOW_MODE
ALLOCATION_WINDOW_DAYS = preallocation_engine.PREALLOCATION_WINDOW_DAYS
ALLOCATION_ETA_LIMIT = date.today() + timedelta(days=ALLOCATION_WINDOW_DAYS)

EXCEL_PATHS = {
    "reservation": BASE_EXCEL_DIR / "Vehicle_Reservation.xlsx",
    "vehicle_tracking": BASE_EXCEL_DIR / "VehicleTracking.xlsx",
    "mc_norm": BASE_EXCEL_DIR / "material code.xlsx",
    "newport": BASE_EXCEL_DIR / "NEWport.xlsx",
    "logistics_db": BASE_EXCEL_DIR / "BASE DE DATOS LOGISTICA.xlsx",
    "priority_orders": BASE_EXCEL_DIR / "quick allocate.xlsx",
    "orders": BASE_EXCEL_DIR / "Dealer Orders.xlsx",
    "not_allocated": BASE_EXCEL_DIR / "Cars not allocated.xlsx",
}

OUTPUT_COLUMNS = [
    "dealer",
    "vin",
    "or_number",
    "budget",
    "e-mail",
    "port",
    "gate_in",
    "eta",
    "p_series",
    "p_trim",
    "int_color",
    "ext_color",
    "client",
    "tag",
    "vessel",
    "country",
]

OR_WITHOUT_CLIENT_OUTPUT_COLUMNS = [
    "dealer",
    "vin",
    "or_number",
    "so_number",
    *OUTPUT_COLUMNS[3:],
]

HEADER_ALIASES = {
    "vin": ["vin", "vin #", "vin#"],
    "dealer": ["dealer", "customer name"],
    "so_number": ["sales order #", "sales order", "so number"],
    "or_number": ["or_number", "or number", "order number", "customer reference #"],
    "budget": ["estimated credit unit", "budget"],
    "sales_company": ["sales company"],
    "port": ["port", "current warehouse"],
    "code": ["material code", "vehicle material code", "code"],
    "p_series": ["product series", "model", "p_series"],
    "p_trim": ["product model", "trim", "p_trim"],
    "int_color": ["interial color", "interior color"],
    "ext_color": ["exterial color", "exterior color"],
    "client": ["account name: account name", "account name", "client"],
}


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


is_missing = tabular.is_missing


def report_progress(message):
    if PROGRESS_CALLBACK is None:
        print(message)
        return

    try:
        PROGRESS_CALLBACK(message)
    except Exception:
        pass


def configure_allocation_window(mode=None, days=None, today=None):
    global ALLOCATION_WINDOW_MODE
    global ALLOCATION_WINDOW_DAYS
    global ALLOCATION_ETA_LIMIT

    selected_mode = mode or ALLOCATION_WINDOW_MODE
    if selected_mode not in ALLOCATION_WINDOW_MODES:
        raise ValueError(
            "allocation window mode must be one of: "
            + ", ".join(sorted(ALLOCATION_WINDOW_MODES))
        )

    selected_days = ALLOCATION_WINDOW_DAYS if days is None else int(days)
    if selected_days < 0:
        raise ValueError("allocation window days must be 0 or greater")

    ALLOCATION_WINDOW_MODE = selected_mode
    ALLOCATION_WINDOW_DAYS = selected_days
    if ALLOCATION_WINDOW_MODE == "gate_in":
        ALLOCATION_ETA_LIMIT = None
    else:
        ALLOCATION_ETA_LIMIT = (today or date.today()) + timedelta(days=ALLOCATION_WINDOW_DAYS)


def allocation_window_label():
    if ALLOCATION_WINDOW_MODE == "gate_in":
        return "gate_in"
    return f"eta_before_{ALLOCATION_WINDOW_DAYS}_days"


format_value = tabular.format_value
normalize_header = tabular.normalize_header
text_key = tabular.text_key
code_key = tabular.code_key
vin_key = tabular.vin_key


def or_number_key(value):
    return format_value(value).upper().replace(" ", "")[:9]


def sales_order_key(value):
    return preallocation_engine.sales_order_key(value)


def priority_reference_matches_reservation(reservation, priority_references):
    """Return whether a reservation matches the mixed Quick Allocate reference list."""
    if isinstance(priority_references, set):
        priority_references = {"or_ow": priority_references, "sales_order": set()}
    return (
        reservation.get("or_number_key", "") in priority_references["or_ow"]
        or reservation.get("so_number_key", "") in priority_references["sales_order"]
    )


def header_index(columns, column_name, required=True):
    return tabular.header_index(columns, HEADER_ALIASES, column_name, required)


def build_indexes(columns, required_columns, optional_columns=None):
    return tabular.build_indexes(columns, HEADER_ALIASES, required_columns, optional_columns)


row_value = tabular.row_value


def open_sheet(key):
    path = EXCEL_PATHS[key]
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = (
        select_active_then_sheet1(workbook, HEADER_ALIASES, ("vin", "port"))
        if key == "not_allocated"
        else workbook.active
    )
    worksheet.reset_dimensions()
    return workbook, worksheet


def read_header(worksheet):
    return [format_value(value) for value in next(worksheet.iter_rows(values_only=True))]


def to_datetime(value):
    if is_missing(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass
    return datetime.fromisoformat(text)


def excel_date_value(value):
    parsed = to_datetime(value)
    return "" if parsed is None else parsed.date()


def parse_positive_budget(value, vin, or_number):
    if is_missing(value):
        return None
    try:
        budget = Decimal(str(value).strip().replace(",", "."))
    except InvalidOperation as exc:
        raise ValueError(f"reservations.budget is not numeric: {vin};{or_number};{value}") from exc
    if not budget.is_finite():
        raise ValueError(f"reservations.budget is not finite: {vin};{or_number};{value}")
    return budget if budget > 0 else None


def budget_output_value(value):
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def reservation_country(sales_company_value):
    return {
        "9030": "SP",
        "9050": "PT",
    }.get(format_value(sales_company_value).upper(), "")


def print_warning(title, rows):
    if not rows:
        return
    print(f"WARNING: {title}: {len(rows)}")
    for row in rows[:20]:
        print(row)


def load_vehicle_tracking():
    old_path = preallocation_engine.EXCEL_PATHS["vehicle_tracking"]
    old_callback = preallocation_engine.PROGRESS_CALLBACK
    try:
        preallocation_engine.EXCEL_PATHS["vehicle_tracking"] = EXCEL_PATHS["vehicle_tracking"]
        preallocation_engine.PROGRESS_CALLBACK = PROGRESS_CALLBACK or report_progress
        return preallocation_engine.load_vehicle_tracking()
    finally:
        preallocation_engine.EXCEL_PATHS["vehicle_tracking"] = old_path
        preallocation_engine.PROGRESS_CALLBACK = old_callback


def clear_vehicle_tracking_cache():
    return preallocation_engine.clear_vehicle_tracking_cache()


def load_mc_norm():
    report_progress("Loading material codes...")
    workbook, worksheet = open_sheet("mc_norm")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["code", "p_series", "p_trim", "int_color", "ext_color"])
        rows = {}
        duplicates = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            code = code_key(row_value(row, indexes, "code"))
            if not code:
                continue
            if code in rows:
                duplicates.add(code)
            rows[code] = {
                "p_series": format_value(row_value(row, indexes, "p_series")),
                "p_trim": format_value(row_value(row, indexes, "p_trim")),
                "int_color": format_value(row_value(row, indexes, "int_color")),
                "ext_color": format_value(row_value(row, indexes, "ext_color")),
            }
        if duplicates:
            raise ValueError("mc_norm duplicated material codes: " + ", ".join(sorted(duplicates)[:20]))
        report_progress(f"Material codes loaded: {len(rows):,}")
        return rows
    finally:
        workbook.close()


def load_newport_ports():
    path = EXCEL_PATHS.get("newport")
    if path is None or not Path(path).exists():
        report_progress("WARNING: NEWport file not found; using the remaining port sources")
        return {}

    report_progress("Loading NEWport ports...")
    workbook, worksheet = open_sheet("newport")
    try:
        columns = read_header(worksheet)
        try:
            indexes = build_indexes(columns, ["vin", "port"])
        except ValueError:
            report_progress(
                "WARNING: NEWport does not contain VIN and port columns; "
                "using the remaining port sources"
            )
            return {}
        ports_by_vin = {}
        duplicated_vins = set()
        required_value_empty = WarningSamples()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row_value(row, indexes, "vin"))
            port = format_value(row_value(row, indexes, "port"))
            if not vin or not port:
                required_value_empty.append(f"{vin};{port}")
                continue
            if vin in ports_by_vin:
                duplicated_vins.add(vin)
            ports_by_vin[vin] = port
        if required_value_empty:
            raise ValueError(
                "NEWport has rows with an empty VIN or port: "
                + ", ".join(required_value_empty[:10])
            )
        if duplicated_vins:
            raise ValueError("NEWport has duplicated VINs: " + ", ".join(sorted(duplicated_vins)[:20]))
        report_progress(f"NEWport VIN ports loaded: {len(ports_by_vin):,}")
        return ports_by_vin
    finally:
        workbook.close()


def load_not_allocated_ports():
    report_progress("Loading Cars not allocated ports...")
    workbook, worksheet = open_sheet("not_allocated")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["vin"], optional_columns=["port"])
        ports_by_vin = {}
        conflicting_vins = WarningSamples()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row_value(row, indexes, "vin"))
            port = format_value(row_value(row, indexes, "port"))
            if not vin or not port:
                continue
            previous_port = ports_by_vin.get(vin)
            if previous_port and text_key(previous_port) != text_key(port):
                conflicting_vins.append(f"{vin};{previous_port};{port}")
                continue
            ports_by_vin.setdefault(vin, port)
        print_warning("Cars not allocated VINs with conflicting ports", conflicting_vins)
        report_progress(f"Cars not allocated VIN ports loaded: {len(ports_by_vin):,}")
        return ports_by_vin
    finally:
        workbook.close()


def load_priority_references():
    report_progress("Loading Quick Allocate references...")
    workbook, worksheet = open_sheet("priority_orders")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["or_number"])
        priority_references = {"or_ow": set(), "sales_order": set()}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            reference_type, key = preallocation_engine.priority_reference(
                row_value(row, indexes, "or_number")
            )
            if reference_type is not None:
                priority_references[reference_type].add(key)
        report_progress(
            "Quick Allocate references loaded: "
            f"{len(priority_references['or_ow']):,} OR/OW and "
            f"{len(priority_references['sales_order']):,} Sales Order"
        )
        return priority_references
    finally:
        workbook.close()


def load_order_clients():
    report_progress("Loading dealer order clients...")
    workbook, worksheet = open_sheet("orders")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["or_number", "client"])
        clients_by_or_number = {}
        conflicting_clients = WarningSamples()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            or_number = or_number_key(row_value(row, indexes, "or_number"))
            if not or_number:
                continue
            client = format_value(row_value(row, indexes, "client"))
            existing_client = clients_by_or_number.get(or_number)
            if existing_client is not None and existing_client != client:
                conflicting_clients.append(f"{or_number};{existing_client};{client}")
                continue
            clients_by_or_number[or_number] = client
        print_warning("dealer orders with conflicting clients", conflicting_clients)
        report_progress(f"Dealer order clients loaded: {len(clients_by_or_number):,}")
        return clients_by_or_number
    finally:
        workbook.close()


def load_reservations():
    report_progress("Loading reservations with positive budget...")
    workbook, worksheet = open_sheet("reservation")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            ["vin", "dealer", "or_number", "budget", "sales_company"],
            ["so_number"],
        )
        rows = []
        summary = Counter()
        required_value_empty = WarningSamples()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            summary["reservation_rows_scanned"] += 1
            sales_company_value = row_value(row, indexes, "sales_company")

            vin = vin_key(row_value(row, indexes, "vin"))
            dealer = format_value(row_value(row, indexes, "dealer"))
            or_number = format_value(row_value(row, indexes, "or_number"))
            if not vin:
                required_value_empty.append(f"{dealer};{or_number}")
                continue
            budget = parse_positive_budget(row_value(row, indexes, "budget"), vin, or_number)
            if budget is None:
                continue
            summary["reservation_rows_positive_budget"] += 1
            country = reservation_country(sales_company_value)
            summary[f"reservation_rows_positive_budget_country_{country or 'UNKNOWN'}"] += 1
            if not or_number:
                summary["reservation_rows_without_or_number"] += 1
            rows.append(
                {
                    "dealer": dealer,
                    "vin": vin,
                    "or_number": or_number,
                    "or_number_key": or_number_key(or_number),
                    "so_number": format_value(row_value(row, indexes, "so_number")),
                    "so_number_key": sales_order_key(row_value(row, indexes, "so_number")),
                    "budget": budget,
                    "country": country,
                }
            )
        print_warning("reservations required value empty", required_value_empty)
        report_progress(f"Reservations with positive budget loaded: {len(rows):,}")
        return rows, summary
    finally:
        workbook.close()


def load_spain_or_reservations_without_client(order_clients):
    report_progress("Loading Spain OR reservations without client...")
    workbook, worksheet = open_sheet("reservation")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            ["vin", "dealer", "or_number", "budget", "sales_company"],
            ["so_number"],
        )
        rows = []
        summary = Counter()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue

            or_number = format_value(row_value(row, indexes, "or_number"))
            or_key = or_number_key(or_number)
            if (
                reservation_country(row_value(row, indexes, "sales_company")) != "SP"
                or not or_key.startswith("OR")
                or format_value(order_clients.get(or_key, ""))
            ):
                continue

            rows.append(
                {
                    "dealer": format_value(row_value(row, indexes, "dealer")),
                    "vin": vin_key(row_value(row, indexes, "vin")),
                    "or_number": or_number,
                    "or_number_key": or_key,
                    "so_number": format_value(row_value(row, indexes, "so_number")),
                    "so_number_key": sales_order_key(row_value(row, indexes, "so_number")),
                    "budget": row_value(row, indexes, "budget"),
                    "country": "SP",
                }
            )

        summary["spain_or_without_client_rows"] = len(rows)
        report_progress(f"Spain OR reservations without client loaded: {len(rows):,}")
        return rows, summary
    finally:
        workbook.close()


def load_port_stock_ports():
    path = EXCEL_PATHS.get("logistics_db")
    if path is None or not Path(path).exists():
        report_progress("WARNING: Logistics database not found; using the existing port sources")
        return {}
    return free_cars_history.load_port_stock_ports(path, report_progress)


def build_rows(
    reservations,
    vehicle_tracking,
    mc_norm,
    newport_ports,
    priority_references,
    order_clients,
    port_stock_ports=None,
    not_allocated_ports=None,
):
    report_progress("Building allocation output...")
    port_stock_ports = port_stock_ports or {}
    not_allocated_ports = not_allocated_ports or {}
    rows = set()
    warnings_by_type = {
        "reservations missing vehicle tracking": WarningSamples(),
        "reservations with empty vehicle tracking ETA": WarningSamples(),
        "reservations with invalid vehicle tracking ETA": WarningSamples(),
        "vehicle tracking material codes missing from material codes": WarningSamples(),
    }
    summary = Counter()
    for reservation in reservations:
        tracking = vehicle_tracking.get(reservation["vin"])
        if tracking is None:
            warnings_by_type["reservations missing vehicle tracking"].append(
                f"{reservation['vin']};{reservation['dealer']};{reservation['or_number']}"
            )
            continue

        eta = tracking.get("eta")
        if ALLOCATION_WINDOW_MODE == "gate_in":
            if is_missing(tracking.get("gate_in")):
                continue
        else:
            if is_missing(eta):
                warnings_by_type["reservations with empty vehicle tracking ETA"].append(
                    f"{reservation['vin']};{reservation['or_number']}"
                )
                continue
            try:
                eta_date = to_datetime(eta).date()
            except ValueError:
                warnings_by_type["reservations with invalid vehicle tracking ETA"].append(
                    f"{reservation['vin']};{reservation['or_number']};{eta}"
                )
                continue
            if eta_date > ALLOCATION_ETA_LIMIT:
                continue

        summary["reservations_eligible_availability_window"] += 1
        material_code = code_key(tracking.get("material_code"))
        mc = mc_norm.get(material_code, {})
        if material_code and not mc:
            warnings_by_type["vehicle tracking material codes missing from material codes"].append(
                f"{reservation['vin']};{material_code}"
            )

        port = resolve_port(
            port_stock_ports.get(reservation["vin"]),
            newport_ports.get(reservation["vin"]),
            not_allocated_ports.get(reservation["vin"]),
            tracking.get("port"),
        )
        rows.add(
            (
                reservation["dealer"],
                reservation["vin"],
                reservation["or_number"],
                budget_output_value(reservation["budget"]),
                "Y" if priority_reference_matches_reservation(reservation, priority_references) else "",
                port,
                excel_date_value(tracking.get("gate_in")),
                excel_date_value(eta),
                mc.get("p_series", ""),
                mc.get("p_trim", ""),
                mc.get("int_color", ""),
                mc.get("ext_color", ""),
                order_clients.get(reservation["or_number_key"], ""),
                format_value(tracking.get("tag")),
                format_value(tracking.get("vessel_name")),
                reservation["country"],
            )
        )

    for title, warning_rows in warnings_by_type.items():
        print_warning(title, warning_rows)

    ordered_rows = sorted(rows, key=lambda row: (text_key(row[0]), row[1], row[2]))
    summary["allocation_rows"] = len(ordered_rows)
    summary["email_y"] = sum(1 for row in ordered_rows if row[4] == "Y")
    for row in ordered_rows:
        country = format_value(row[15]).upper() or "UNKNOWN"
        summary[f"allocation_rows_country_{country}"] += 1
    report_progress(f"Allocation rows built: {len(ordered_rows):,}")
    return ordered_rows, summary


def build_spain_or_without_client_rows(
    reservations,
    vehicle_tracking,
    mc_norm,
    newport_ports,
    priority_references,
    port_stock_ports=None,
    not_allocated_ports=None,
):
    report_progress("Building Spain OR reservations without client output...")
    port_stock_ports = port_stock_ports or {}
    not_allocated_ports = not_allocated_ports or {}
    rows = set()
    for reservation in reservations:
        tracking = vehicle_tracking.get(reservation["vin"], {})
        material_code = code_key(tracking.get("material_code"))
        mc = mc_norm.get(material_code, {})
        port = resolve_port(
            port_stock_ports.get(reservation["vin"]),
            newport_ports.get(reservation["vin"]),
            not_allocated_ports.get(reservation["vin"]),
            tracking.get("port"),
        )
        rows.add(
            (
                reservation["dealer"],
                reservation["vin"],
                reservation["or_number"],
                reservation.get("so_number", ""),
                reservation["budget"],
                "Y" if priority_reference_matches_reservation(reservation, priority_references) else "",
                port,
                excel_date_value(tracking.get("gate_in")),
                excel_date_value(tracking.get("eta")),
                mc.get("p_series", ""),
                mc.get("p_trim", ""),
                mc.get("int_color", ""),
                mc.get("ext_color", ""),
                "",
                format_value(tracking.get("tag")),
                format_value(tracking.get("vessel_name")),
                reservation["country"],
            )
        )

    ordered_rows = sorted(rows, key=lambda row: (text_key(row[0]), row[1], row[2]))
    report_progress(f"Spain OR reservations without client rows built: {len(ordered_rows):,}")
    return ordered_rows


def write_output(rows, summary, spain_or_without_client_rows=None):
    report_progress("Writing allocation workbook...")
    workbook = Workbook(write_only=True)
    try:
        worksheet = workbook.create_sheet("ALLOCATION")
        row_values = lambda row: list(row)
        column_widths = calculate_column_widths(OUTPUT_COLUMNS, rows, row_values, format_value)
        prepare_worksheet(worksheet, OUTPUT_COLUMNS, column_widths, 40)
        for row in rows:
            append_row(worksheet, row_values(row), SHORT_DATE_FORMAT)

        without_client_rows = spain_or_without_client_rows or []
        worksheet = workbook.create_sheet("OR WITHOUT CLIENT")
        column_widths = calculate_column_widths(
            OR_WITHOUT_CLIENT_OUTPUT_COLUMNS,
            without_client_rows,
            row_values,
            format_value,
        )
        prepare_worksheet(worksheet, OR_WITHOUT_CLIENT_OUTPUT_COLUMNS, column_widths, 40)
        for row in without_client_rows:
            append_row(worksheet, row_values(row), SHORT_DATE_FORMAT)
        save_workbook_atomically(workbook, OUTPUT_XLSX_PATH)
    finally:
        workbook.close()

    print("PROCESS_SUMMARY")
    for key in (
        "reservation_rows_scanned",
        "reservation_rows_positive_budget",
        "reservation_rows_without_or_number",
        "reservations_eligible_availability_window",
        "allocation_rows",
        "spain_or_without_client_rows",
        "email_y",
    ):
        print(f"{key}: {summary[key]}")
    print_country_summary(summary, "reservation_rows_positive_budget_country_", "reservations_positive_budget_by_country")
    print_country_summary(summary, "allocation_rows_country_", "allocation_rows_by_country")
    print(f"availability_window: {allocation_window_label()}")
    print(f"OUTPUT_XLSX: {OUTPUT_XLSX_PATH}")


def print_country_summary(summary, key_prefix, label):
    country_counts = {
        key.removeprefix(key_prefix): value
        for key, value in summary.items()
        if key.startswith(key_prefix)
    }
    values = "; ".join(
        f"{country}={country_counts[country]}"
        for country in sorted(country_counts)
    )
    print(f"{label}: {values or 'none'}")


def main():
    report_progress("Starting vehicle allocation...")
    mc_norm = load_mc_norm()
    vehicle_tracking = load_vehicle_tracking()
    newport_ports = load_newport_ports()
    not_allocated_ports = load_not_allocated_ports()
    port_stock_ports = load_port_stock_ports()
    priority_references = load_priority_references()
    order_clients = load_order_clients()
    reservations, summary = load_reservations()
    spain_or_without_client_reservations, without_client_summary = load_spain_or_reservations_without_client(
        order_clients
    )
    rows, row_summary = build_rows(
        reservations,
        vehicle_tracking,
        mc_norm,
        newport_ports,
        priority_references,
        order_clients,
        port_stock_ports,
        not_allocated_ports,
    )
    summary.update(row_summary)
    summary.update(without_client_summary)
    spain_or_without_client_rows = build_spain_or_without_client_rows(
        spain_or_without_client_reservations,
        vehicle_tracking,
        mc_norm,
        newport_ports,
        priority_references,
        port_stock_ports,
        not_allocated_ports,
    )
    summary["spain_or_without_client_rows"] = len(spain_or_without_client_rows)
    write_output(rows, summary, spain_or_without_client_rows)


if __name__ == "__main__":
    main()

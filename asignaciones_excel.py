from collections import Counter
import csv
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
import sys
import unicodedata
import warnings

from openpyxl import Workbook, load_workbook

from excel_output import append_row, calculate_column_widths, prepare_worksheet, save_workbook_atomically
import free_cars_history
from port_resolution import resolve_port as resolve_operational_port
import vehicle_tracking_cache as tracking_cache
from warning_samples import WarningSamples


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

BASE_EXCEL_DIR = Path.cwd()
OUTPUT_XLSX_PATH = BASE_EXCEL_DIR / "asignaciones_resultado.xlsx"
OUTPUT_DELIMITER = ";"
SHORT_DATE_FORMAT = "yyyy-mm-dd"
CREATE_TIME_MIN = None
PREALLOCATION_WINDOW_MODE = "eta_days"
PREALLOCATION_WINDOW_DAYS = 7
NON_OR_ETA_BEFORE_LIMIT = datetime.combine(date.today() + timedelta(days=7), time.min)
RESERVATION_ETA_LIMIT = datetime.combine(date.today() + timedelta(days=7), time.max)
PROGRESS_CALLBACK = None
SHIPPING_ETA_OVERRIDES = {}
CACHE_DIR = tracking_cache.default_cache_dir()
VEHICLE_TRACKING_CACHE_PATH = CACHE_DIR / "vehicle_tracking.pkl"
VEHICLE_TRACKING_CACHE_VERSION = tracking_cache.CACHE_VERSION

PREALLOCATION_WINDOW_MODES = {"gate_in", "eta_days"}

EXCEL_PATHS = {
    "sototal": BASE_EXCEL_DIR / "sototal.xlsx",
    "mc_norm": BASE_EXCEL_DIR / "material code.xlsx",
    "dealer_info": BASE_EXCEL_DIR / "Dealer Info.xlsx",
    "not_allocated": BASE_EXCEL_DIR / "Cars not allocated.xlsx",
    "vehicle_tracking": BASE_EXCEL_DIR / "VehicleTracking.xlsx",
    "newport": BASE_EXCEL_DIR / "NEWport.xlsx",
    "logistics_db": BASE_EXCEL_DIR / "BASE DE DATOS LOGISTICA.xlsx",
    "reservations": BASE_EXCEL_DIR / "Vehicle_Reservation.xlsx",
    "priority_orders": BASE_EXCEL_DIR / "quick allocate.xlsx",
}
SHEET_NAMES = {
    "not_allocated": "Sheet1",
}

FIXED_MATERIAL_CODE_SALES_ORDERS = set()
FIXED_MATERIAL_CODE_SALES_ORDER_KEYS = {
    str(sales_order).strip().upper()
    for sales_order in FIXED_MATERIAL_CODE_SALES_ORDERS
}

STEP4_BLOCKED_MODEL_RULES = {}
# Kept for compatibility with older standalone entry points.
STEP4_BLOCKED_P_SERIES = set(STEP4_BLOCKED_MODEL_RULES)
STEP4_BLOCKED_P_SERIES_KEYS = {
    str(p_series).strip().upper()
    for p_series in STEP4_BLOCKED_P_SERIES
}

EXCLUDED_CREATED_BY_USERS = set()
EXCLUDED_CREATED_BY_KEYS = {
    re.sub(r"[^A-Z0-9]", "", str(user).strip().upper())
    for user in EXCLUDED_CREATED_BY_USERS
}

EXCLUDED_DSN_CONTAINS = set()
EXCLUDED_DSN_CONTAINS_KEYS = {
    str(dsn_text).strip().upper()
    for dsn_text in EXCLUDED_DSN_CONTAINS
    if str(dsn_text).strip()
}

MODEL_YEAR_CUT_RULES = []

HEADER_ALIASES = {
    "code": ["material code", "vehicle material code", "code"],
    "p_series": ["product series", "p_series"],
    "product_model": ["product model", "trim"],
    "int_color": ["interial color", "interior color"],
    "ext_color": ["exterial color", "exterior color"],
    "model_year": ["model year", "model_year", "my"],
    "sales_order": ["so #", "sales order #", "sales_order", "sales order"],
    "or_number": ["customer reference #", "customer reference", "or number", "or_number"],
    "create_time": ["create time", "create_time"],
    "material_code": ["vehicle material code", "material code", "material_code"],
    "country": ["country"],
    "sap": ["customer#", "customer# (sap code)", "customer# \uff08sap code\uff09", "sap"],
    "dsn": ["delivery short name", "dsn"],
    "qty": ["ordered qty", "qty"],
    "reserved_with": ["reserve with vin qty", "reserved with vin qty"],
    "bo_with": ["bo with vin qty"],
    "bo_without": ["bo without vin qty"],
    "dn_qty": ["dn qty"],
    "gate_out_qty": ["gate out qty"],
    "created_by": ["created by"],
    "match_group": ["match group", "match_group"],
    "blocked_groups": ["blocked groups", "blocked_groups"],
    "package": ["package"],
    "dealer": ["dealer", "customer name"],
    "vin": ["vin #", "vin"],
    "note": ["note"],
    "port": ["port", "current warehouse"],
    "description": ["description(local)", "description"],
    "vessel": ["name of vessel", "vessel"],
    "priority": ["priority", "\u4f18\u5148\u5206\u914d"],
    "gate_in": ["gate in date", "gate in", "gate_in"],
    "gate_out": ["gate out date", "actual gate out"],
    "production_date": ["production date"],
    "eta": ["purchase eta", "eta"],
    "status": ["current status", "status"],
    "invoice_date": ["sales invoice date"],
    "customer_country": ["customer country"],
    "address": ["detail delivery address", "address"],
    "city": ["city"],
    "tag": ["tag name", "tag"],
    "related_order": ["related order#", "related order", "related_order"],
    "reserved_so": ["reserved so #", "reserved so", "reserved_so"],
    "dn_create_time": ["dn create time", "dn_create_time"],
    "allocation_date": ["allocation date", "allocate date", "allocation_date"],
    "sales_company": ["sales company"],
    "budget": ["estimated credit unit", "budget"],
}


def is_missing(value):
    return value is None or str(value).strip() == ""


def report_progress(message):
    if PROGRESS_CALLBACK is None:
        return

    try:
        PROGRESS_CALLBACK(message)
    except Exception:
        pass


def configure_preallocation_window(mode=None, days=None, today=None):
    global PREALLOCATION_WINDOW_MODE
    global PREALLOCATION_WINDOW_DAYS
    global NON_OR_ETA_BEFORE_LIMIT
    global RESERVATION_ETA_LIMIT

    selected_mode = mode or PREALLOCATION_WINDOW_MODE
    if selected_mode not in PREALLOCATION_WINDOW_MODES:
        raise ValueError(
            "preallocation window mode must be one of: "
            + ", ".join(sorted(PREALLOCATION_WINDOW_MODES))
        )

    selected_days = PREALLOCATION_WINDOW_DAYS if days is None else int(days)
    if selected_mode == "eta_days" and selected_days < 0:
        raise ValueError("preallocation window days must be 0 or greater")

    run_date = today or date.today()
    PREALLOCATION_WINDOW_MODE = selected_mode
    PREALLOCATION_WINDOW_DAYS = selected_days

    if PREALLOCATION_WINDOW_MODE == "gate_in":
        NON_OR_ETA_BEFORE_LIMIT = None
        RESERVATION_ETA_LIMIT = None
        return

    NON_OR_ETA_BEFORE_LIMIT = datetime.combine(
        run_date + timedelta(days=PREALLOCATION_WINDOW_DAYS),
        time.min,
    )
    RESERVATION_ETA_LIMIT = datetime.combine(
        run_date + timedelta(days=PREALLOCATION_WINDOW_DAYS),
        time.max,
    )


def preallocation_window_label():
    if PREALLOCATION_WINDOW_MODE == "gate_in":
        return "gate_in"

    return f"eta_before_{PREALLOCATION_WINDOW_DAYS}_days"


def vehicle_tracking_file_signature():
    return tracking_cache.vehicle_tracking_file_signature(EXCEL_PATHS["vehicle_tracking"])


def clear_vehicle_tracking_cache():
    return tracking_cache.clear_cache(VEHICLE_TRACKING_CACHE_PATH)


def load_vehicle_tracking_cache(signature):
    return tracking_cache.load_cache(VEHICLE_TRACKING_CACHE_PATH, signature, report_progress)


def write_vehicle_tracking_cache(signature, by_vin):
    tracking_cache.write_cache(VEHICLE_TRACKING_CACHE_PATH, signature, by_vin, report_progress)


def format_value(value):
    if is_missing(value):
        return ""

    return str(value).replace("\r", " ").replace("\n", " ").strip()


def normalize_header(value):
    value = unicodedata.normalize("NFKC", format_value(value))
    value = value.lower()
    value = re.sub(r"\s+", " ", value)
    return value.replace("-", " ").replace("_", " ").strip()


def header_index(columns, column_name, required=True):
    aliases = {normalize_header(alias) for alias in HEADER_ALIASES[column_name]}
    for index, column in enumerate(columns):
        if normalize_header(column) in aliases:
            return index

    if not required:
        return None

    raise ValueError(f"Missing column {column_name}. Headers: {columns}")


def build_indexes(columns, required_columns, optional_columns=None):
    indexes = {column: header_index(columns, column) for column in required_columns}
    for column in optional_columns or []:
        indexes[column] = header_index(columns, column, required=False)
    return indexes


def max_required_col(indexes):
    return max(index for index in indexes.values() if index is not None) + 1


def row_value(row, indexes, column):
    index = indexes[column]
    if index is None or index >= len(row):
        return None

    return row[index]


def text_key(value):
    value = format_value(value).upper()
    value = unicodedata.normalize("NFD", value)
    value = "".join(char for char in value if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", value).strip()


def code_key(value):
    return format_value(value).upper()


def identifier_key(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return format_value(value).upper()


def sales_order_key(value):
    return identifier_key(value)


def vin_key(value):
    return code_key(value)


def created_by_key(value):
    return format_value(value).upper()


def created_by_exclusion_key(value):
    return re.sub(r"[^A-Z0-9]", "", created_by_key(value))


def is_excluded_created_by(value):
    return created_by_exclusion_key(value) in EXCLUDED_CREATED_BY_KEYS


def is_excluded_dsn(value):
    dsn = format_value(value).upper()
    return any(text in dsn for text in EXCLUDED_DSN_CONTAINS_KEYS)


def or_number_priority_key(value):
    return format_value(value).upper().replace(" ", "")[:9]


def to_number(value):
    if is_missing(value):
        return 0
    if isinstance(value, (int, float)):
        return value

    return float(str(value).strip().replace(",", "."))


def to_datetime(value):
    if is_missing(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    for date_format in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass

    return datetime.fromisoformat(text)


def as_datetime(value):
    parsed = to_datetime(value)
    if parsed is None:
        raise ValueError(f"Date value is empty: {value}")

    return parsed


def open_sheet(table_name, sheet_name=None):
    path = EXCEL_PATHS[table_name]
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    selected_sheet = sheet_name or SHEET_NAMES.get(table_name)
    if selected_sheet and selected_sheet in workbook.sheetnames:
        worksheet = workbook[selected_sheet]
    else:
        worksheet = workbook.active

    worksheet.reset_dimensions()
    return worksheet


def close_sheet_workbook(worksheet):
    workbook = getattr(worksheet, "parent", None)
    if workbook is not None:
        workbook.close()


def read_header(worksheet):
    return [format_value(value) for value in next(worksheet.iter_rows(values_only=True))]


def passes_sototal_cleaning(row, indexes):
    available_qty = (
        to_number(row_value(row, indexes, "qty"))
        - to_number(row_value(row, indexes, "reserved_with"))
        - to_number(row_value(row, indexes, "bo_with"))
        - to_number(row_value(row, indexes, "dn_qty"))
        - to_number(row_value(row, indexes, "gate_out_qty"))
    )
    create_time = to_datetime(row_value(row, indexes, "create_time"))
    return (
        available_qty > 0
        and to_number(row_value(row, indexes, "bo_without")) > 0
        and create_time is not None
        and (CREATE_TIME_MIN is None or create_time >= CREATE_TIME_MIN)
    )


def print_warning(title, rows):
    if not rows:
        return

    print(f"WARNING: {title}: {len(rows)}")
    for row in rows[:20]:
        print(row)


def parse_budget(value):
    try:
        budget = Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError(f"reservations.budget is not numeric: {value}") from exc

    if not budget.is_finite():
        raise ValueError(f"reservations.budget is not finite: {value}")

    return budget


def collapse_dealer_candidates(candidates):
    collapsed = {}
    for candidate in candidates:
        if is_missing(candidate.get("package")):
            continue
        business_key = (
            text_key(candidate["dealer"]),
            candidate["package"],
            candidate.get("match_group", ""),
            candidate.get("blocked_groups", ""),
        )
        collapsed.setdefault(business_key, candidate)

    return list(collapsed.values())


def load_mc_norm():
    report_progress("Reading material codes...")
    worksheet = open_sheet("mc_norm")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["code", "p_series", "product_model", "int_color", "ext_color", "model_year"])
        by_code = {}
        code_counter = Counter()
        skipped = WarningSamples()

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(not is_missing(value) for value in row):
                continue

            code = code_key(row_value(row, indexes, "code"))
            p_series = format_value(row_value(row, indexes, "p_series"))
            product_model = format_value(row_value(row, indexes, "product_model"))
            int_color = format_value(row_value(row, indexes, "int_color"))
            ext_color = format_value(row_value(row, indexes, "ext_color"))
            model_year = format_value(row_value(row, indexes, "model_year"))

            if any(is_missing(value) for value in [code, p_series, product_model, int_color, ext_color, model_year]):
                skipped.append((code, "required value is empty"))
                continue

            code_counter[code] += 1
            by_code[code] = {
                "code": code,
                "p_series": p_series,
                "p_trim": product_model,
                "aggr": f"{p_series}{product_model}{int_color}{ext_color}",
                "model_year": model_year,
            }

        duplicated = sorted(code for code, count in code_counter.items() if count > 1)
        if duplicated:
            raise ValueError(f"mc_norm.code duplicated values: {', '.join(duplicated[:10])}")

        print_warning("mc_norm rows skipped", skipped)
        report_progress(f"Material codes loaded: {len(by_code):,}")
        return by_code
    finally:
        close_sheet_workbook(worksheet)


def load_vehicle_tracking():
    signature = vehicle_tracking_file_signature()
    cached = load_vehicle_tracking_cache(signature)
    if cached is not None:
        return tracking_cache.apply_shipping_eta_overrides(
            cached,
            SHIPPING_ETA_OVERRIDES,
            report_progress,
        )

    report_progress("Reading VehicleTracking.xlsx...")
    worksheet = open_sheet("vehicle_tracking")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            ["vin", "eta", "port", "gate_in", "status"],
            optional_columns=[
                "material_code",
                "description",
                "vessel",
                "dsn",
                "sap",
                "gate_out",
                "production_date",
                "invoice_date",
                "customer_country",
                "address",
                "city",
                "tag",
                "related_order",
                "reserved_so",
                "dn_create_time",
                "allocation_date",
            ],
        )
        max_col = max_required_col(indexes)
        report_progress(f"VehicleTracking.xlsx columns limited to {max_col} of {len(columns)}")
        by_vin = {}
        vin_counter = Counter()
        scanned_rows = 0

        for scanned_rows, row in enumerate(worksheet.iter_rows(min_row=2, max_col=max_col, values_only=True), start=1):
            if scanned_rows % 25000 == 0:
                report_progress(
                    f"Reading VehicleTracking.xlsx: {scanned_rows:,} rows scanned, {len(by_vin):,} VINs loaded"
                )
            if not any(not is_missing(value) for value in row):
                continue

            vin = vin_key(row_value(row, indexes, "vin"))
            if is_missing(vin):
                continue

            vin_counter[vin] += 1
            by_vin[vin] = {
                "vin": vin,
                "material_code": code_key(row_value(row, indexes, "material_code")),
                "description": format_value(row_value(row, indexes, "description")),
                "eta": row_value(row, indexes, "eta"),
                "port": format_value(row_value(row, indexes, "port")),
                "vessel_name": format_value(row_value(row, indexes, "vessel")),
                "dsn": format_value(row_value(row, indexes, "dsn")),
                "sap": code_key(row_value(row, indexes, "sap")),
                "gate_in": row_value(row, indexes, "gate_in"),
                "gate_out": row_value(row, indexes, "gate_out"),
                "production_date": row_value(row, indexes, "production_date"),
                "status": format_value(row_value(row, indexes, "status")),
                "invoice_date": row_value(row, indexes, "invoice_date"),
                "country": format_value(row_value(row, indexes, "customer_country")),
                "address": format_value(row_value(row, indexes, "address")),
                "city": format_value(row_value(row, indexes, "city")),
                "tag": format_value(row_value(row, indexes, "tag")),
                "related_order": format_value(row_value(row, indexes, "related_order")),
                "reserved_so": format_value(row_value(row, indexes, "reserved_so")),
                "dn_create_time": row_value(row, indexes, "dn_create_time"),
                "allocation_date": row_value(row, indexes, "allocation_date"),
            }

        duplicated = sorted(vin for vin, count in vin_counter.items() if count > 1)
        if duplicated:
            raise ValueError(f"vehicle_tracking.vin duplicated values: {', '.join(duplicated[:10])}")

        report_progress(f"Vehicle tracking loaded: {len(by_vin):,} VINs from {scanned_rows:,} rows")
        write_vehicle_tracking_cache(signature, by_vin)
        return tracking_cache.apply_shipping_eta_overrides(
            by_vin,
            SHIPPING_ETA_OVERRIDES,
            report_progress,
        )
    finally:
        close_sheet_workbook(worksheet)


def load_newport_ports():
    path = EXCEL_PATHS.get("newport")
    if path is None or not Path(path).exists():
        report_progress("WARNING: NEWport file not found; using the remaining port sources")
        return {}

    report_progress("Reading NEWport.xlsx...")
    worksheet = open_sheet("newport")
    try:
        columns = read_header(worksheet)
        try:
            indexes = build_indexes(columns, ["vin", "port"])
        except ValueError:
            report_progress(
                "WARNING: NEWport does not contain VIN and port columns; "
                "using the remaining port sources"
            )
            return {}
        ports_by_vin = {}
        duplicated_vins = set()
        required_value_empty = []

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(not is_missing(value) for value in row):
                continue

            vin = vin_key(row_value(row, indexes, "vin"))
            port = format_value(row_value(row, indexes, "port"))
            if is_missing(vin) or is_missing(port):
                required_value_empty.append(f"{vin};{port}")
                continue
            if vin in ports_by_vin:
                duplicated_vins.add(vin)
            ports_by_vin[vin] = port

        if required_value_empty:
            examples = ", ".join(required_value_empty[:10])
            raise ValueError(
                "NEWport has rows with an empty VIN or port: "
                f"{len(required_value_empty)}. Examples: {examples}"
            )
        if duplicated_vins:
            raise ValueError(
                "NEWport has duplicated VINs: " + ", ".join(sorted(duplicated_vins)[:10])
            )

        report_progress(f"NEWport ports loaded: {len(ports_by_vin):,}")
        return ports_by_vin
    finally:
        close_sheet_workbook(worksheet)


def load_port_stock_ports():
    path = EXCEL_PATHS.get("logistics_db")
    if path is None or not Path(path).exists():
        report_progress("WARNING: Logistics database not found; using the existing port sources")
        return {}
    return free_cars_history.load_port_stock_ports(path, report_progress)


def resolve_port(port_stock_port, newport_port, not_allocated_port, vehicle_tracking_port):
    return resolve_operational_port(
        port_stock_port,
        newport_port,
        not_allocated_port,
        vehicle_tracking_port,
    )


def load_dealer_info():
    report_progress("Reading dealer info...")
    worksheet = open_sheet("dealer_info")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            ["dsn", "sap", "match_group", "blocked_groups", "package", "dealer"],
        )
        by_dsn_sap = {}

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(not is_missing(value) for value in row):
                continue

            dsn = format_value(row_value(row, indexes, "dsn"))
            sap = identifier_key(row_value(row, indexes, "sap"))
            dealer_row = {
                "dsn": dsn,
                "dsn_key": text_key(dsn),
                "sap": sap,
                "match_group": format_match_groups(
                    parse_match_groups(
                        row_value(row, indexes, "match_group"),
                        f"Dealer info match_group row {row_number}",
                    )
                ),
                "blocked_groups": format_match_groups(
                    parse_match_groups(
                        row_value(row, indexes, "blocked_groups"),
                        f"Dealer info blocked_groups row {row_number}",
                    )
                ),
                "package": format_value(row_value(row, indexes, "package")),
                "dealer": format_value(row_value(row, indexes, "dealer")),
            }
            if dealer_row["dsn_key"] and sap:
                by_dsn_sap.setdefault((dealer_row["dsn_key"], sap), []).append(dealer_row)

        report_progress(f"Dealer info loaded: {sum(len(rows) for rows in by_dsn_sap.values()):,} rows")
        return by_dsn_sap
    finally:
        close_sheet_workbook(worksheet)


def load_orders(mc_norm, dealer_info):
    report_progress("Reading and cleaning orders...")
    worksheet = open_sheet("sototal")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            [
                "sales_order",
                "or_number",
                "create_time",
                "material_code",
                "country",
                "sap",
                "dsn",
                "qty",
                "reserved_with",
                "bo_with",
                "bo_without",
                "dn_qty",
                "gate_out_qty",
                "created_by",
            ],
        )
        orders = []
        warnings_by_type = {
            "missing_mc_norm": WarningSamples(),
            "missing_dealer": WarningSamples(),
            "duplicate_dealer_matches": WarningSamples(),
            "required_value_empty": WarningSamples(),
        }
        summary = Counter()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue

            dsn = format_value(row_value(row, indexes, "dsn"))
            if is_excluded_dsn(dsn):
                summary["excluded"] += 1
                continue
            if not passes_sototal_cleaning(row, indexes):
                continue

            created_by = format_value(row_value(row, indexes, "created_by"))
            if is_excluded_created_by(created_by):
                summary["excluded"] += 1
                continue

            sales_order = identifier_key(row_value(row, indexes, "sales_order"))
            or_number = format_value(row_value(row, indexes, "or_number"))
            material_code = code_key(row_value(row, indexes, "material_code"))
            country = format_value(row_value(row, indexes, "country")).upper()
            sap = identifier_key(row_value(row, indexes, "sap"))
            create_time = to_datetime(row_value(row, indexes, "create_time"))

            mc = mc_norm.get(material_code)
            if mc is None:
                warnings_by_type["missing_mc_norm"].append(f"{sales_order};{material_code}")
                continue

            dealer_candidates = collapse_dealer_candidates(dealer_info.get((text_key(dsn), sap), []))
            if not dealer_candidates:
                warnings_by_type["missing_dealer"].append(f"{sales_order};{sap};{dsn}")
                continue
            if len(dealer_candidates) > 1:
                warnings_by_type["duplicate_dealer_matches"].append(
                    f"{sales_order};{sap};{dsn};{len(dealer_candidates)}"
                )
                continue

            dealer = dealer_candidates[0]
            order = (
                sales_order,
                or_number,
                create_time,
                material_code,
                country,
                mc["aggr"],
                mc["p_series"],
                mc["p_trim"],
                dsn,
                dealer["match_group"],
                dealer["blocked_groups"],
                dealer["dsn"],
                dealer["package"],
                mc["model_year"],
                dealer["dealer"],
            )
            if any(is_missing(value) for value in [sales_order, create_time, material_code, mc["aggr"], mc["p_series"], dsn, dealer["dsn"], dealer["package"], dealer["dealer"]]):
                warnings_by_type["required_value_empty"].append(f"{sales_order};{sap};{dsn}")
                continue

            orders.append(order)

        for title, rows in warnings_by_type.items():
            print_warning(f"orders {title}", rows)

        report_progress(f"Orders loaded after cleaning: {len(orders):,}")
        return orders, summary
    finally:
        close_sheet_workbook(worksheet)


def is_portugal_reserved_note(note):
    normalized_note = text_key(note)
    if normalized_note == "PT RESERVATION INTERNAL SPONSORSHIP":
        return False

    return (
        "PORTUGAL" in normalized_note
        or "PT RESERVATION" in normalized_note
        or "PT WAREHOUSE" in normalized_note
    )


def load_available_cars(
    mc_norm,
    vehicle_tracking,
    newport_ports,
    port_stock_ports=None,
    include_portugal_reserved_cars=False,
):
    report_progress("Reading cars not allocated...")
    port_stock_ports = port_stock_ports or {}
    worksheet = open_sheet("not_allocated", "Sheet1")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            ["vin", "material_code", "note", "port", "priority", "match_group", "gate_in"],
            ["description"],
        )
        cars = []
        portugal_reserved_cars = []
        warnings_by_type = {
            "missing_mc_norm": WarningSamples(),
            "required_value_empty": WarningSamples(),
            "unexpected_priority": WarningSamples(),
        }
        summary = Counter()
        seen_vins = set()
        duplicated_vins = set()
        scanned_rows = 0

        for scanned_rows, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
            if scanned_rows % 5000 == 0:
                report_progress(
                    f"Reading cars not allocated: {scanned_rows:,} rows scanned, {len(cars):,} cars loaded"
                )
            if not any(not is_missing(value) for value in row):
                continue

            vin = vin_key(row_value(row, indexes, "vin"))
            not_allocated_material_code = code_key(row_value(row, indexes, "material_code"))
            note = format_value(row_value(row, indexes, "note"))
            na_port = format_value(row_value(row, indexes, "port"))
            na_gate_in = row_value(row, indexes, "gate_in")

            if include_portugal_reserved_cars and vin and not_allocated_material_code and is_portugal_reserved_note(note):
                portugal_reserved_cars.append(
                    {
                        "vin": vin,
                        "material_code": not_allocated_material_code,
                        "description": format_value(row_value(row, indexes, "description")),
                        "note": note,
                        "gate_in": na_gate_in,
                    }
                )

            if note:
                continue

            priority = format_value(row_value(row, indexes, "priority"))
            match_group = format_match_groups(
                parse_match_groups(
                    row_value(row, indexes, "match_group"),
                    f"Cars not allocated Match group row {scanned_rows + 1}",
                )
            )
            if is_priority_excluded(priority) and not match_group:
                summary["priority_zero_excluded"] += 1
                continue
            if is_priority_excluded(priority):
                summary["priority_zero_group_only"] += 1

            if (
                not is_missing(priority)
                and not is_priority_excluded(priority)
                and priority_rank(priority) is None
            ):
                warnings_by_type["unexpected_priority"].append(f"{vin};{priority}")

            tracking = vehicle_tracking.get(vin)
            if tracking is None:
                summary["missing_vehicle_tracking"] += 1
                continue

            material_code = code_key(tracking.get("material_code")) or not_allocated_material_code

            status = format_value(tracking["status"]).upper()
            if status in {"", "OFFLINE"}:
                summary["offline_or_empty_status"] += 1
                continue

            mc = mc_norm.get(material_code)
            if mc is None:
                warnings_by_type["missing_mc_norm"].append(f"{vin};{material_code}")
                continue

            eta = tracking["eta"] if not is_missing(tracking["eta"]) else na_gate_in
            port_stock_port = port_stock_ports.get(vin, "")
            newport_port = newport_ports.get(vin, "")
            port = resolve_port(port_stock_port, newport_port, na_port, tracking["port"])
            if port_stock_port:
                summary["port_from_port_stock"] += 1
            elif newport_port:
                summary["port_from_newport"] += 1
            car = (
                vin,
                material_code,
                tracking["gate_in"],
                eta,
                mc["aggr"],
                port,
                priority,
                match_group,
                mc["model_year"],
            )
            if any(is_missing(value) for value in [vin, material_code, eta, mc["aggr"], port]):
                warnings_by_type["required_value_empty"].append(f"{vin};{material_code}")
                continue

            if vin in seen_vins:
                duplicated_vins.add(vin)
            seen_vins.add(vin)
            cars.append(car)

        if duplicated_vins:
            raise ValueError(f"available_cars.vin duplicated values: {', '.join(sorted(duplicated_vins)[:10])}")

        for title, rows in warnings_by_type.items():
            print_warning(f"available_cars {title}", rows)

        report_progress(f"Available cars loaded: {len(cars):,} from {scanned_rows:,} rows")
        if include_portugal_reserved_cars:
            return cars, summary, portugal_reserved_cars
        return cars, summary
    finally:
        close_sheet_workbook(worksheet)


def load_priority_keys():
    report_progress("Reading priority orders...")
    worksheet = open_sheet("priority_orders")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["or_number"])
        keys = set()
        key_counter = Counter()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue

            priority_key = or_number_priority_key(row_value(row, indexes, "or_number"))
            if is_missing(priority_key):
                continue
            keys.add(priority_key)
            key_counter[priority_key] += 1

        duplicated = sorted(key for key, count in key_counter.items() if count > 1)
        print_warning("priority duplicate OR/OW keys", [f"{key};{key_counter[key]}" for key in duplicated])
        report_progress(f"Priority OR/OW keys loaded: {len(keys):,}")
        return keys
    finally:
        close_sheet_workbook(worksheet)


def load_reservations(vehicle_tracking):
    report_progress("Reading reservations...")
    worksheet = open_sheet("reservations")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["sales_company", "budget", "dealer", "vin"])
        reservations = []
        warnings_by_type = {
            "missing_vehicle_tracking": WarningSamples(),
            "tracking_eta_empty": WarningSamples(),
            "required_value_empty": WarningSamples(),
        }

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue

            sales_company = identifier_key(row_value(row, indexes, "sales_company"))
            if sales_company != "9030":
                continue

            budget = row_value(row, indexes, "budget")
            dealer = format_value(row_value(row, indexes, "dealer"))
            vin = vin_key(row_value(row, indexes, "vin"))
            if any(is_missing(value) for value in [budget, dealer, vin]):
                warnings_by_type["required_value_empty"].append(f"{sales_company};{dealer};{vin}")
                continue

            tracking = vehicle_tracking.get(vin)
            if tracking is None:
                warnings_by_type["missing_vehicle_tracking"].append(f"{sales_company};{dealer};{vin}")
                continue

            if PREALLOCATION_WINDOW_MODE == "gate_in":
                if is_missing(tracking["gate_in"]):
                    continue
            else:
                if is_missing(tracking["eta"]):
                    warnings_by_type["tracking_eta_empty"].append(f"{sales_company};{dealer};{vin}")
                    continue
                if as_datetime(tracking["eta"]) > RESERVATION_ETA_LIMIT:
                    continue

            reservations.append((budget, dealer, vin))

        for title, rows in warnings_by_type.items():
            print_warning(f"reservations {title}", rows)

        report_progress(f"Reservations loaded: {len(reservations):,}")
        return reservations
    finally:
        close_sheet_workbook(worksheet)


def build_reservation_budget_state(reservations, reservation_columns):
    budget_index = reservation_columns.index("budget")
    dealer_index = reservation_columns.index("dealer")
    vin_index = reservation_columns.index("vin")
    min_budget_by_dealer = {}
    reserved_vins_by_dealer = {}

    for row in reservations:
        dealer = text_key(row[dealer_index])
        budget = parse_budget(row[budget_index])
        vin = vin_key(row[vin_index])
        if dealer not in min_budget_by_dealer or budget < min_budget_by_dealer[dealer]:
            min_budget_by_dealer[dealer] = budget
        reserved_vins_by_dealer.setdefault(dealer, set()).add(vin)

    reserved_amount_by_dealer = Counter(
        {dealer: len(vins) for dealer, vins in reserved_vins_by_dealer.items()}
    )
    return min_budget_by_dealer, reserved_amount_by_dealer


def warn_reservation_dealers_without_orders(orders, order_columns, reservations, reservation_columns):
    order_dealer_index = order_columns.index("dealer")
    reservation_dealer_index = reservation_columns.index("dealer")
    order_dealers = {
        text_key(row[order_dealer_index])
        for row in orders
        if not is_missing(row[order_dealer_index])
    }
    missing = {}

    for row in reservations:
        dealer = row[reservation_dealer_index]
        key = text_key(dealer)
        if key not in order_dealers:
            missing.setdefault(key, format_value(dealer))

    print_warning("reservation dealers without matching orders.dealer", sorted(missing.values()))


def priority_rank(value):
    text = format_value(value)
    if not text:
        return None

    try:
        rank = Decimal(text)
    except InvalidOperation:
        return None

    if not rank.is_finite() or rank < 1 or rank != rank.to_integral_value():
        return None

    return int(rank)


def is_priority_excluded(value):
    text = format_value(value)
    if not text:
        return False

    try:
        priority = Decimal(text)
    except InvalidOperation:
        return False

    return priority.is_finite() and priority == 0


def parse_match_groups(value, field_name):
    text = format_value(value)
    if not text:
        return ()

    groups = []
    seen = set()
    for raw_group in text.split(";"):
        raw_group = raw_group.strip()
        if not raw_group:
            raise ValueError(f"{field_name} contains an empty match group: {text}")

        group = raw_group.upper()
        if not re.fullmatch(r"[A-Z]+", group):
            try:
                group_number = Decimal(raw_group)
            except InvalidOperation as exc:
                raise ValueError(f"{field_name} has an invalid match group: {raw_group}") from exc
            if (
                not group_number.is_finite()
                or group_number < 1
                or group_number != group_number.to_integral_value()
            ):
                raise ValueError(f"{field_name} has an invalid match group: {raw_group}")
            group = match_group_letters(int(group_number))
        if group not in seen:
            groups.append(group)
            seen.add(group)

    return tuple(groups)


def match_group_letters(number):
    """Convert legacy positive numeric groups to Excel-style letters: 1 -> A."""
    letters = []
    while number:
        number, remainder = divmod(number - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def format_match_groups(groups):
    return ";".join(groups)


def order_match_groups(order, order_columns):
    if "match_group" not in order_columns:
        return ()
    return parse_match_groups(order[order_columns.index("match_group")], "orders.match_group")


def order_blocked_groups(order, order_columns):
    if "blocked_groups" not in order_columns:
        return ()
    return parse_match_groups(order[order_columns.index("blocked_groups")], "orders.blocked_groups")


def car_match_groups(car, car_columns):
    if "match_group" not in car_columns:
        return ()
    return parse_match_groups(car[car_columns.index("match_group")], "available_cars.match_group")


def match_group_route_for_order(order, car, order_columns, car_columns):
    order_groups = order_match_groups(order, order_columns)
    blocked_groups = set(order_blocked_groups(order, order_columns))
    car_groups = car_match_groups(car, car_columns)
    if blocked_groups.intersection(car_groups):
        return None
    common_groups = [group for group in order_groups if group in car_groups]
    priority_zero = is_priority_excluded(car[car_columns.index("priority")])

    if priority_zero:
        if not car_groups or not common_groups:
            return None
        return "exclusive group", order_groups.index(common_groups[0])

    if not car_groups:
        return "general pool", len(order_groups)
    if common_groups:
        return "preferred group", order_groups.index(common_groups[0])
    return "other group", len(order_groups)


def match_group_candidate_key(order, car, order_columns, car_columns, original_index):
    route = match_group_route_for_order(order, car, order_columns, car_columns)
    if route is None:
        return (2, float("inf"), float("inf"), 2, datetime.max, 1, original_index)

    route_name, group_position = route
    priority = priority_rank(car[car_columns.index("priority")])
    if is_priority_excluded(car[car_columns.index("priority")]):
        priority_key = (0, 0)
    elif priority is not None:
        priority_key = (1, priority)
    else:
        priority_key = (2, float("inf"))

    if is_car_in_port(car, car_columns):
        availability_key = (0, sort_datetime_or_max(car[car_columns.index("gate_in")]))
    else:
        availability_key = (1, sort_datetime_or_max(car[car_columns.index("eta")]))

    if route_name in {"preferred group", "exclusive group"}:
        # Direct dealer-group matches always precede the shared remainder pool.
        return (0, group_position, *priority_key, *availability_key, 0, original_index)

    # General cars and cars from another group share the same pool. Group membership
    # is only the final tie-breaker after priority and availability.
    other_group_key = 1 if route_name == "other group" else 0
    return (
        1,
        float("inf"),
        *priority_key,
        *availability_key,
        other_group_key,
        original_index,
    )


def is_car_in_port(car, car_columns):
    return not is_missing(car[car_columns.index("gate_in")])


def sort_datetime_or_max(value):
    parsed = to_datetime(value)
    return parsed if parsed is not None else datetime.max


def can_use_material_code(order, car, order_columns, car_columns):
    sales_order = sales_order_key(order[order_columns.index("sales_order")])
    if sales_order not in FIXED_MATERIAL_CODE_SALES_ORDER_KEYS:
        return True

    order_material_code = code_key(order[order_columns.index("material_code")])
    car_material_code = code_key(car[car_columns.index("material_code")])
    return order_material_code == car_material_code


def can_use_dealer_budget(order, order_columns, dealer_budget_by_dealer, dealer_reserved_amount_by_dealer):
    if dealer_budget_by_dealer is None:
        return True

    country = format_value(order[order_columns.index("country")]).upper()
    if country in {"PT", "PORTUGAL"}:
        return True

    dealer = text_key(order[order_columns.index("dealer")])
    if dealer not in dealer_budget_by_dealer:
        return True

    return dealer_budget_by_dealer[dealer] > Decimal(dealer_reserved_amount_by_dealer[dealer])


def can_use_eta_before_limit(car, car_columns, eta_before_limit):
    if eta_before_limit is None:
        return True

    eta = car[car_columns.index("eta")]
    if is_missing(eta):
        return False

    return as_datetime(eta) < eta_before_limit


def can_use_preallocation_window(car, car_columns, eta_before_limit):
    if PREALLOCATION_WINDOW_MODE == "gate_in":
        return is_car_in_port(car, car_columns)

    return can_use_eta_before_limit(car, car_columns, eta_before_limit)


def consume_dealer_budget(order, order_columns, dealer_budget_by_dealer, dealer_reserved_amount_by_dealer):
    if dealer_budget_by_dealer is None:
        return

    dealer = text_key(order[order_columns.index("dealer")])
    if dealer in dealer_budget_by_dealer:
        dealer_reserved_amount_by_dealer[dealer] += 1


def normalize_model_year(value):
    if is_missing(value):
        return ""

    value = format_value(value).upper().replace(" ", "")
    if value.startswith("MY"):
        return value
    if re.fullmatch(r"20\d{2}(\.\d+)?", value):
        return "MY" + value[2:]
    if re.fullmatch(r"\d{2}(\.\d+)?", value):
        return "MY" + value

    return value


def model_year_number(value):
    normalized = normalize_model_year(value)
    match = re.fullmatch(r"MY(\d{2})(\.\d+)?", normalized)
    if not match:
        return None

    return Decimal(match.group(1) + (match.group(2) or ""))


def is_same_side_of_model_year_cut(order_model_year, car_model_year, min_new_model_year):
    order_year = model_year_number(order_model_year)
    car_year = model_year_number(car_model_year)
    if order_year is None or car_year is None:
        return True

    return (order_year >= min_new_model_year) == (car_year >= min_new_model_year)


def can_use_model_year(order, car, order_columns, car_columns):
    order_p_series = format_value(order[order_columns.index("p_series")]).upper()
    order_aggr = format_value(order[order_columns.index("aggr")]).upper()

    for rule in MODEL_YEAR_CUT_RULES:
        if order_p_series != rule["p_series"]:
            continue
        if rule["aggr_contains"] not in order_aggr:
            continue

        return is_same_side_of_model_year_cut(
            order[order_columns.index("model_year")],
            car[car_columns.index("model_year")],
            rule["min_new_model_year"],
        )

    return True


def aggr_key(value):
    return format_value(value).upper()


def is_or_order(order, order_columns):
    return "OR" in format_value(order[order_columns.index("or_number")]).upper()


def is_not_or_order(order, order_columns):
    return not is_or_order(order, order_columns)


def is_step4_blocked_model(order, order_columns):
    p_series_key = text_key(order[order_columns.index("p_series")])
    exceptions = STEP4_BLOCKED_MODEL_RULES.get(p_series_key)
    if exceptions is None:
        return False

    p_trim = ""
    if "p_trim" in order_columns:
        p_trim = text_key(order[order_columns.index("p_trim")])

    return not any(exception in p_trim for exception in exceptions)


def sort_orders_for_match(orders_to_sort, order_columns):
    return sorted(
        orders_to_sort,
        key=lambda row: (
            aggr_key(row[order_columns.index("aggr")]),
            row[order_columns.index("create_time")],
            row[order_columns.index("sales_order")],
        ),
    )


def sort_cars_for_match(cars, car_columns):
    def car_sort_key(row):
        vin = row[car_columns.index("vin")]
        if is_car_in_port(row, car_columns):
            priority = priority_rank(row[car_columns.index("priority")])
            return (
                0,
                priority is None,
                priority if priority is not None else float("inf"),
                sort_datetime_or_max(row[car_columns.index("gate_in")]),
                vin,
            )

        return (
            1,
            True,
            float("inf"),
            sort_datetime_or_max(row[car_columns.index("eta")]),
            vin,
        )

    return sorted(
        cars,
        key=car_sort_key,
    )


def match_orders(
    orders_to_match,
    available_by_aggr,
    order_columns,
    car_columns,
    use_preallocation_window=False,
    eta_before_limit=None,
    dealer_budget_by_dealer=None,
    dealer_reserved_amount_by_dealer=None,
):
    matches = []
    unmatched_orders = []
    material_mismatches = []

    for order in orders_to_match:
        if not can_use_dealer_budget(
            order,
            order_columns,
            dealer_budget_by_dealer,
            dealer_reserved_amount_by_dealer,
        ):
            unmatched_orders.append(order)
            continue

        order_aggr = order[order_columns.index("aggr")]
        candidates = available_by_aggr.get(aggr_key(order_aggr), [])
        match_position = None
        candidates_by_preference = sorted(
            enumerate(candidates),
            key=lambda item: match_group_candidate_key(
                order,
                item[1],
                order_columns,
                car_columns,
                item[0],
            ),
        )

        for index, car in candidates_by_preference:
            if match_group_route_for_order(order, car, order_columns, car_columns) is None:
                continue
            if use_preallocation_window and not can_use_preallocation_window(car, car_columns, eta_before_limit):
                continue
            if not can_use_material_code(order, car, order_columns, car_columns):
                continue
            if not can_use_model_year(order, car, order_columns, car_columns):
                continue
            match_position = index
            break

        if match_position is None:
            unmatched_orders.append(order)
            continue

        car = candidates.pop(match_position)
        match_group_route, _ = match_group_route_for_order(
            order,
            car,
            order_columns,
            car_columns,
        )
        consume_dealer_budget(
            order,
            order_columns,
            dealer_budget_by_dealer,
            dealer_reserved_amount_by_dealer,
        )

        order_material_code = order[order_columns.index("material_code")]
        car_material_code = car[car_columns.index("material_code")]
        if order_material_code != car_material_code:
            material_mismatches.append(
                (
                    order[order_columns.index("sales_order")],
                    order_material_code,
                    car_material_code,
                    order[order_columns.index("package")],
                    order[order_columns.index("or_number")],
                    order_aggr,
                    order[order_columns.index("model_year")],
                    car[car_columns.index("aggr")],
                    car[car_columns.index("model_year")],
                    car[car_columns.index("eta")],
                )
            )

        matches.append(
            (
                order[order_columns.index("sales_order")],
                car[car_columns.index("vin")],
                order[order_columns.index("or_number")],
                order[order_columns.index("p_series")],
                car[car_columns.index("eta")],
                car[car_columns.index("port")],
                car[car_columns.index("aggr")],
                order[order_columns.index("create_time")],
                order[order_columns.index("dsn")] if "dsn" in order_columns else "",
                order[order_columns.index("country")],
                format_match_groups(order_match_groups(order, order_columns)),
                format_match_groups(car_match_groups(car, car_columns)),
                match_group_route,
                car[car_columns.index("priority")],
            )
        )

    return matches, unmatched_orders, material_mismatches


def excel_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    return value


def write_sheet(workbook, title, headers, rows):
    worksheet = workbook.create_sheet(title)
    row_values = lambda row: [excel_value(value) for value in row]
    column_widths = calculate_column_widths(headers, rows, row_values, format_value)
    prepare_worksheet(worksheet, headers, column_widths, 45)
    for row in rows:
        append_row(worksheet, row_values(row), SHORT_DATE_FORMAT)


def build_portugal_reserved_car_rows(orders, order_columns, portugal_reserved_cars):
    sales_order_index = order_columns.index("sales_order")
    material_code_index = order_columns.index("material_code")
    country_index = order_columns.index("country")
    create_time_index = order_columns.index("create_time")

    cars_by_material_code = {}
    for car in portugal_reserved_cars:
        cars_by_material_code.setdefault(code_key(car["material_code"]), []).append(car)

    for cars in cars_by_material_code.values():
        cars.sort(
            key=lambda car: (
                text_key(car["note"]),
                sort_datetime_or_max(car["gate_in"]),
                car["vin"],
            )
        )

    portugal_orders = sorted(
        (
            order
            for order in orders
            if text_key(order[country_index]) in {"PT", "PORTUGAL"}
        ),
        key=lambda order: (
            sort_datetime_or_max(order[create_time_index]),
            code_key(order[material_code_index]),
            order[sales_order_index],
        ),
    )

    rows = []
    for order in portugal_orders:
        sales_order = order[sales_order_index]
        material_code = code_key(order[material_code_index])
        for car in cars_by_material_code.get(material_code, []):
            rows.append(
                (
                    sales_order,
                    car["vin"],
                    car["material_code"],
                    car["description"],
                    car["note"],
                )
            )

    return rows


def write_output_workbook(material_mismatches, final_matches, portugal_reserved_cars=None):
    report_progress("Writing output workbook...")
    portugal_reserved_cars = portugal_reserved_cars or []
    workbook = Workbook(write_only=True)
    try:
        write_sheet(
            workbook,
            "MATERIAL_CODE_MISMATCHES",
            [
                "sales_order",
                "order_material_code",
                "car_material_code",
                "package",
                "or_number",
                "order_aggr",
                "order_model_year",
                "car_aggr",
                "car_model_year",
                "eta",
            ],
            material_mismatches,
        )
        write_sheet(
            workbook,
            "FINAL_MATCHES",
            [
                "sales_order",
                "vin",
                "or_number",
                "p_series",
                "eta",
                "port",
                "aggr",
                "create_time",
                "dsn",
                "country",
                "dealer_match_group",
                "car_match_group",
                "match_group_route",
                "priority",
            ],
            final_matches,
        )
        write_sheet(
            workbook,
            "PORTUGAL_RESERVED_CARS",
            ["sales_order", "vin", "material_code", "description", "note"],
            portugal_reserved_cars,
        )
        save_workbook_atomically(workbook, OUTPUT_XLSX_PATH)
    finally:
        workbook.close()
    report_progress(f"Output workbook written: {OUTPUT_XLSX_PATH}")


def print_outputs(step_results, order_groups, all_matches, portugal_reserved_cars=None):
    output_writer = csv.writer(sys.stdout, delimiter=OUTPUT_DELIMITER, lineterminator="\n")
    matched_country_counts = Counter(format_value(row[9]).upper() for row in all_matches)
    matched_total = len(all_matches)
    sp_matches = matched_country_counts["SPAIN"] + matched_country_counts["SP"]
    pt_matches = matched_country_counts["PORTUGAL"] + matched_country_counts["PT"]
    other_matches = matched_total - sp_matches - pt_matches
    sp_pct = sp_matches / matched_total * 100 if matched_total else 0
    pt_pct = pt_matches / matched_total * 100 if matched_total else 0
    other_pct = other_matches / matched_total * 100 if matched_total else 0

    print()
    print("PROCESS_SUMMARY")
    for index, (matches, unmatched, mismatches) in enumerate(step_results, start=1):
        print(f"step{index}_orders: {len(order_groups[index - 1])}")
        print(f"step{index}_matches: {len(matches)}")
        print(f"step{index}_unmatched: {len(unmatched)}")
        print(f"step{index}_material_mismatches: {len(mismatches)}")
    print(
        "country_sp_pt_proportion: "
        f"SP {sp_matches}/{matched_total} ({sp_pct:.1f}%) | "
        f"PT {pt_matches}/{matched_total} ({pt_pct:.1f}%) | "
        f"other {other_matches}/{matched_total} ({other_pct:.1f}%)"
    )

    print()
    print("MATERIAL_CODE_MISMATCHES")
    output_writer.writerow(
        [
            "sales_order",
            "order_material_code",
            "car_material_code",
            "package",
            "or_number",
            "order_aggr",
            "order_model_year",
            "car_aggr",
            "car_model_year",
            "eta",
        ]
    )
    all_material_mismatches = []
    for _, _, mismatches in step_results:
        all_material_mismatches.extend(mismatches)
    output_writer.writerows(all_material_mismatches)

    print()
    print("FINAL_MATCHES")
    output_writer.writerow(
        [
            "sales_order",
            "vin",
            "or_number",
            "p_series",
            "eta",
            "port",
            "aggr",
            "create_time",
            "dsn",
            "country",
            "dealer_match_group",
            "car_match_group",
            "match_group_route",
            "priority",
        ]
    )
    output_writer.writerows(all_matches)

    write_output_workbook(all_material_mismatches, all_matches, portugal_reserved_cars)
    print()
    print(f"OUTPUT_XLSX: {OUTPUT_XLSX_PATH}")


def main():
    report_progress("Starting preallocation...")
    mc_norm = load_mc_norm()
    vehicle_tracking = load_vehicle_tracking()
    newport_ports = load_newport_ports()
    port_stock_ports = load_port_stock_ports()
    dealer_info = load_dealer_info()
    orders, orders_summary = load_orders(mc_norm, dealer_info)
    available_cars, available_cars_summary, portugal_reserved_cars = load_available_cars(
        mc_norm,
        vehicle_tracking,
        newport_ports,
        port_stock_ports,
        include_portugal_reserved_cars=True,
    )
    priority_keys = load_priority_keys()
    reservations = load_reservations(vehicle_tracking)

    print()
    print("INPUT_SUMMARY")
    print(f"orders_loaded: {len(orders)}")
    print(f"orders_excluded: {orders_summary['excluded']}")
    print(f"available_cars_loaded: {len(available_cars)}")
    print(f"port_stock_ports_loaded: {len(port_stock_ports)}")
    print(f"available_cars_port_from_port_stock: {available_cars_summary['port_from_port_stock']}")
    print(f"newport_ports_loaded: {len(newport_ports)}")
    print(f"available_cars_port_from_newport: {available_cars_summary['port_from_newport']}")
    print(f"available_cars_missing_vehicle_tracking: {available_cars_summary['missing_vehicle_tracking']}")
    print(f"available_cars_offline_or_empty_status: {available_cars_summary['offline_or_empty_status']}")
    print(f"available_cars_priority_zero_excluded: {available_cars_summary['priority_zero_excluded']}")
    print(f"available_cars_priority_zero_group_only: {available_cars_summary['priority_zero_group_only']}")
    print(f"priority_or_ow_keys_loaded: {len(priority_keys)}")
    print(f"reservations_loaded: {len(reservations)}")
    print(f"preallocation_window: {preallocation_window_label()}")

    order_columns = [
        "sales_order",
        "or_number",
        "create_time",
        "material_code",
        "country",
        "aggr",
        "p_series",
        "p_trim",
        "dsn",
        "match_group",
        "blocked_groups",
        "dealer_dsn",
        "package",
        "model_year",
        "dealer",
    ]
    car_columns = [
        "vin",
        "material_code",
        "gate_in",
        "eta",
        "aggr",
        "port",
        "priority",
        "match_group",
        "model_year",
    ]
    reservation_columns = ["budget", "dealer", "vin"]
    portugal_reserved_car_rows = build_portugal_reserved_car_rows(
        orders,
        order_columns,
        portugal_reserved_cars,
    )

    report_progress("Checking reservation dealers...")
    warn_reservation_dealers_without_orders(orders, order_columns, reservations, reservation_columns)
    report_progress("Building reservation budget state...")
    reservation_budget_by_dealer, reservation_reserved_amount_by_dealer = build_reservation_budget_state(
        reservations=reservations,
        reservation_columns=reservation_columns,
    )

    order_priority_keys = {
        or_number_priority_key(row[order_columns.index("or_number")])
        for row in orders
        if not is_missing(row[order_columns.index("or_number")])
    }
    priority_not_in_orders = sorted(priority_keys - order_priority_keys)
    print_warning("priority OR/OW keys not in clean orders", priority_not_in_orders)

    orders_with_priority_columns = order_columns + ["priority"]
    orders_with_priority = [
        row + ("Y" if or_number_priority_key(row[order_columns.index("or_number")]) in priority_keys else "N",)
        for row in orders
    ]

    blocked_model_priority_or_orders = [
        row
        for row in orders_with_priority
        if row[-1] == "Y"
        and is_or_order(row, orders_with_priority_columns)
        and is_step4_blocked_model(row, orders_with_priority_columns)
    ]
    if blocked_model_priority_or_orders:
        print_warning(
            "Blocked-model OR orders marked as priority; continuing as priority",
            [
                (
                    f"{row[orders_with_priority_columns.index('sales_order')]};"
                    f"{row[orders_with_priority_columns.index('p_series')]};"
                    f"{row[orders_with_priority_columns.index('p_trim')]};"
                    f"{row[orders_with_priority_columns.index('or_number')]}"
                )
                for row in blocked_model_priority_or_orders
            ],
        )

    report_progress("Preparing available-car pools...")
    available_by_aggr = {}
    for car in sort_cars_for_match(available_cars, car_columns):
        aggr = aggr_key(car[car_columns.index("aggr")])
        available_by_aggr.setdefault(aggr, []).append(car)

    priority_or_orders = sort_orders_for_match(
        [
            row
            for row in orders_with_priority
            if row[-1] == "Y"
            and is_or_order(row, orders_with_priority_columns)
        ],
        orders_with_priority_columns,
    )
    or_orders = sort_orders_for_match(
        [
            row
            for row in orders_with_priority
            if row[-1] == "N"
            and is_or_order(row, orders_with_priority_columns)
        ],
        orders_with_priority_columns,
    )
    priority_not_or_orders = sort_orders_for_match(
        [
            row
            for row in orders_with_priority
            if row[-1] == "Y"
            and is_not_or_order(row, orders_with_priority_columns)
        ],
        orders_with_priority_columns,
    )
    not_priority_not_or_orders = sort_orders_for_match(
        [
            row
            for row in orders_with_priority
            if row[-1] == "N"
            and is_not_or_order(row, orders_with_priority_columns)
            and not is_step4_blocked_model(row, orders_with_priority_columns)
        ],
        orders_with_priority_columns,
    )

    report_progress(f"Step 1: matching priority OR orders ({len(priority_or_orders):,})...")
    step1 = match_orders(
        orders_to_match=priority_or_orders,
        available_by_aggr=available_by_aggr,
        order_columns=orders_with_priority_columns,
        car_columns=car_columns,
    )
    report_progress(f"Step 1 complete: {len(step1[0]):,} matches")
    report_progress(f"Step 2: matching non-priority OR orders ({len(or_orders):,})...")
    step2 = match_orders(
        orders_to_match=or_orders,
        available_by_aggr=available_by_aggr,
        order_columns=orders_with_priority_columns,
        car_columns=car_columns,
    )
    report_progress(f"Step 2 complete: {len(step2[0]):,} matches")
    report_progress(f"Step 3: matching priority non-OR orders ({len(priority_not_or_orders):,})...")
    step3 = match_orders(
        orders_to_match=priority_not_or_orders,
        available_by_aggr=available_by_aggr,
        order_columns=orders_with_priority_columns,
        car_columns=car_columns,
        use_preallocation_window=True,
        eta_before_limit=NON_OR_ETA_BEFORE_LIMIT,
    )
    report_progress(f"Step 3 complete: {len(step3[0]):,} matches")
    report_progress(f"Step 4: matching non-priority non-OR orders ({len(not_priority_not_or_orders):,})...")
    step4 = match_orders(
        orders_to_match=not_priority_not_or_orders,
        available_by_aggr=available_by_aggr,
        order_columns=orders_with_priority_columns,
        car_columns=car_columns,
        use_preallocation_window=True,
        eta_before_limit=NON_OR_ETA_BEFORE_LIMIT,
        dealer_budget_by_dealer=reservation_budget_by_dealer,
        dealer_reserved_amount_by_dealer=reservation_reserved_amount_by_dealer,
    )
    report_progress(f"Step 4 complete: {len(step4[0]):,} matches")

    all_matches = step1[0] + step2[0] + step3[0] + step4[0]
    report_progress("Checking duplicates...")
    duplicated_sales_orders = sorted(
        value for value, count in Counter(row[0] for row in all_matches).items() if count > 1
    )
    duplicated_vins = sorted(
        value for value, count in Counter(row[1] for row in all_matches).items() if count > 1
    )
    if duplicated_sales_orders:
        raise ValueError(f"Duplicated matched sales_order values: {', '.join(duplicated_sales_orders)}")
    if duplicated_vins:
        raise ValueError(f"Duplicated matched vin values: {', '.join(duplicated_vins)}")

    print_outputs(
        step_results=[step1, step2, step3, step4],
        order_groups=[priority_or_orders, or_orders, priority_not_or_orders, not_priority_not_or_orders],
        all_matches=all_matches,
        portugal_reserved_cars=portugal_reserved_car_rows,
    )


if __name__ == "__main__":
    main()

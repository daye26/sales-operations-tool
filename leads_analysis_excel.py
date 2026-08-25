"""Build the Spain and Portugal daily Leads analysis workbook."""

import ast
import csv
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
import unicodedata
import warnings

from openpyxl import Workbook, load_workbook

from excel_output import append_row, calculate_column_widths, prepare_worksheet, save_workbook_atomically


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

BASE_EXCEL_DIR = Path.cwd()
OUTPUT_XLSX_PATH = BASE_EXCEL_DIR / "leads_analysis_result.xlsx"
PROGRESS_CALLBACK = None
SHORT_DATE_FORMAT = "yyyy-mm-dd"

EXCEL_PATHS = {
    "leads_sp": BASE_EXCEL_DIR / "Leads_SP.csv",
    "leads_pt": BASE_EXCEL_DIR / "Leads_PT.csv",
    "model_eq": BASE_EXCEL_DIR / "model_eq.xlsx",
}

LEAD_HEADER_ALIASES = {
    "create_date": ["create date"],
    "request_type": ["request type"],
    "interest_model": ["interest model"],
    "interest_model_external": ["interest model (external)"],
}
MODEL_EQ_HEADER_ALIASES = {
    "id": ["id"],
    "model": ["model"],
}
METRICS = (
    ("offer", "offer_request"),
    ("test drive request", "test_drive_request"),
    ("test drive", "test_drive"),
    ("dealer leads", "dealer_leads"),
)
DEFAULT_TEST_DRIVE_FORMULA = "test_drive_request + dealer_leads * 0.3"
TEST_DRIVE_FORMULA_VARIABLES = (
    "test_drive_request",
    "dealer_leads",
    "offer_request",
)


def report_progress(message):
    if PROGRESS_CALLBACK is None:
        print(message)
        return
    try:
        PROGRESS_CALLBACK(message)
    except Exception:
        pass


def is_missing(value):
    return value is None or str(value).strip() == ""


def format_value(value):
    if is_missing(value):
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def normalize_header(value):
    text = unicodedata.normalize("NFKC", format_value(value)).lower()
    text = text.replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", text).strip()


def text_key(value):
    return format_value(value).upper()


def parse_test_drive_formula(value):
    """Parse the deliberately small arithmetic language for the derived metric."""
    formula = format_value(value)
    if not formula:
        raise ValueError("Test drive formula cannot be empty.")

    try:
        expression = ast.parse(formula, mode="eval").body
    except SyntaxError as exc:
        raise ValueError("Invalid test drive formula. Use the allowed variables and +, -, *, /, ( ).") from exc

    def validate(node):
        if isinstance(node, ast.Name):
            if node.id not in TEST_DRIVE_FORMULA_VARIABLES:
                allowed = ", ".join(TEST_DRIVE_FORMULA_VARIABLES)
                raise ValueError(f"Unknown test drive variable {node.id!r}. Allowed variables: {allowed}.")
            return
        if isinstance(node, ast.Constant):
            if isinstance(node.value, bool) or not isinstance(node.value, (int, float)):
                raise ValueError("Test drive formula only accepts numeric constants.")
            return
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            validate(node.operand)
            return
        if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
            validate(node.left)
            validate(node.right)
            return
        raise ValueError("Test drive formula only supports +, -, *, / and parentheses.")

    validate(expression)
    return formula, expression


def evaluate_test_drive_formula(expression, bucket):
    if isinstance(expression, ast.Name):
        return Decimal(str(bucket[expression.id]))
    if isinstance(expression, ast.Constant):
        return Decimal(str(expression.value))
    if isinstance(expression, ast.UnaryOp):
        value = evaluate_test_drive_formula(expression.operand, bucket)
        return value if isinstance(expression.op, ast.UAdd) else -value

    left = evaluate_test_drive_formula(expression.left, bucket)
    right = evaluate_test_drive_formula(expression.right, bucket)
    if isinstance(expression.op, ast.Add):
        return left + right
    if isinstance(expression.op, ast.Sub):
        return left - right
    if isinstance(expression.op, ast.Mult):
        return left * right
    if right == 0:
        # Zero is exported as a blank metric cell, matching the other no-activity values.
        return Decimal(0)
    return left / right


def calculate_metric_value(bucket, metric_key, test_drive_expression):
    if metric_key != "test_drive":
        return bucket[metric_key]

    value = evaluate_test_drive_formula(test_drive_expression, bucket)
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def header_index(headers, aliases, source_label, field):
    normalized_aliases = {normalize_header(alias) for alias in aliases}
    for index, header in enumerate(headers):
        if normalize_header(header) in normalized_aliases:
            return index
    raise ValueError(f"{source_label} is missing column {field}. Headers: {headers}")


def row_value(row, index):
    return row[index] if index < len(row) else None


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = format_value(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    for date_format in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            pass
    return None


def parse_period_date(value):
    parsed = parse_date(value)
    if parsed is None:
        raise ValueError(f"Invalid Leads date: {value!r}. Use dd/mm/yyyy.")
    return parsed


def read_tabular_source(source_path):
    source_path = Path(source_path)
    if not source_path.exists():
        raise FileNotFoundError(source_path)

    if source_path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            worksheet.reset_dimensions()
            try:
                headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
            except StopIteration as exc:
                raise ValueError(f"{source_path.name} is empty.") from exc
            rows = [tuple(row) for row in worksheet.iter_rows(min_row=2, values_only=True)]
            return headers, rows
        finally:
            workbook.close()

    if source_path.suffix.lower() != ".csv":
        raise ValueError(f"Unsupported Leads source: {source_path.name}. Use CSV or Excel.")

    decode_error = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with source_path.open("r", encoding=encoding, newline="") as source_file:
                sample = source_file.read(8192)
                source_file.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(source_file, dialect)
                try:
                    headers = next(reader)
                except StopIteration as exc:
                    raise ValueError(f"{source_path.name} is empty.") from exc
                return headers, [tuple(row) for row in reader]
        except UnicodeDecodeError as exc:
            decode_error = exc
    raise ValueError(f"Could not decode {source_path.name} as CSV.") from decode_error


def load_leads(country):
    source_key = f"leads_{country.lower()}"
    source_path = Path(EXCEL_PATHS[source_key])
    source_label = f"Leads {country}"
    report_progress(f"Reading {source_label}...")
    headers, rows = read_tabular_source(source_path)
    indexes = {
        field: header_index(headers, aliases, source_label, field)
        for field, aliases in LEAD_HEADER_ALIASES.items()
    }

    records = []
    id_counts = Counter()
    invalid_dates = []
    skipped_without_id = 0
    for row in rows:
        if not any(not is_missing(value) for value in row):
            continue
        interest_model = text_key(row_value(row, indexes["interest_model"]))
        external_model = text_key(row_value(row, indexes["interest_model_external"]))
        lead_id = interest_model or external_model
        if not lead_id:
            skipped_without_id += 1
            continue

        id_counts[lead_id] += 1
        create_date_value = row_value(row, indexes["create_date"])
        create_date = parse_date(create_date_value)
        if create_date is None:
            invalid_dates.append(lead_id)
            continue

        request_type = text_key(row_value(row, indexes["request_type"])) or "DEALER LEADS"
        records.append(
            {
                "country": country,
                "create_date": create_date,
                "request_type": request_type,
                "id": lead_id,
            }
        )

    report_progress(f"{source_label} loaded: {len(records):,} valid dated rows")
    return {
        "records": records,
        "id_counts": id_counts,
        "rows_scanned": len(rows),
        "skipped_without_id": skipped_without_id,
        "invalid_dates": invalid_dates,
    }


def load_model_equivalences():
    source_path = Path(EXCEL_PATHS["model_eq"])
    report_progress("Reading Model equivalence...")
    headers, rows = read_tabular_source(source_path)
    indexes = {
        field: header_index(headers, aliases, "Model equivalence", field)
        for field, aliases in MODEL_EQ_HEADER_ALIASES.items()
    }

    mapping = {}
    duplicate_ids = 0
    for row in rows:
        if not any(not is_missing(value) for value in row):
            continue
        lead_id = text_key(row_value(row, indexes["id"]))
        model = format_value(row_value(row, indexes["model"]))
        if not lead_id or not model:
            continue
        existing_model = mapping.get(lead_id)
        if existing_model is not None:
            if text_key(existing_model) != text_key(model):
                raise ValueError(
                    "Model equivalence maps ID "
                    f"{lead_id!r} to more than one model: {existing_model!r}; {model!r}"
                )
            duplicate_ids += 1
            continue
        mapping[lead_id] = model

    report_progress(f"Model equivalence loaded: {len(mapping):,} IDs")
    return mapping, duplicate_ids


def build_reporting_models(model_equivalences):
    """Return the unique non-OLD model catalogue used by each country report."""
    models_by_key = {}
    for model in model_equivalences.values():
        normalized_model = format_value(model)
        model_key = text_key(normalized_model)
        if not model_key or model_key == "OLD":
            continue
        models_by_key.setdefault(model_key, normalized_model)
    return [models_by_key[model_key] for model_key in sorted(models_by_key)]


def build_metric_rows(counts, dates, report_models=None, test_drive_expression=None):
    if test_drive_expression is None:
        _, test_drive_expression = parse_test_drive_formula(DEFAULT_TEST_DRIVE_FORMULA)

    ordered_dates = sorted(dates)
    headers = ["model", "metric", *(day.isoformat() for day in ordered_dates)]
    rows = []
    models_by_key = {}
    for model in report_models or ():
        model_key = text_key(model)
        if model_key:
            models_by_key.setdefault(model_key, format_value(model))
    for model, _ in counts:
        model_key = text_key(model)
        if model_key:
            models_by_key.setdefault(model_key, format_value(model))
    models = [models_by_key[model_key] for model_key in sorted(models_by_key)]
    metrics = sorted(METRICS, key=lambda metric: metric[0])
    for metric_index, (metric_label, metric_key) in enumerate(metrics):
        if metric_index:
            rows.append(tuple(None for _ in headers))
        for model in models:
            row = [model, metric_label]
            for day in ordered_dates:
                bucket = counts[(model, day)]
                value = calculate_metric_value(bucket, metric_key, test_drive_expression)
                row.append(value or None)
            rows.append(tuple(row))
    return rows, headers


def build_old_rows(old_counts_by_country, old_dates_by_country, test_drive_expression=None):
    if test_drive_expression is None:
        _, test_drive_expression = parse_test_drive_formula(DEFAULT_TEST_DRIVE_FORMULA)

    ordered_dates = sorted({day for dates in old_dates_by_country.values() for day in dates})
    headers = ["model", "country", "metric", *(day.isoformat() for day in ordered_dates)]
    models = sorted(
        {
            model
            for counts in old_counts_by_country.values()
            for model, _ in counts
        },
        key=text_key,
    )
    rows = []
    for model in models:
        countries = [
            country
            for country in ("SP", "PT")
            if any(stored_model == model for stored_model, _ in old_counts_by_country[country])
        ]
        for metric_label, metric_key in METRICS:
            for country in countries:
                row = [model, country, metric_label]
                for day in ordered_dates:
                    bucket = old_counts_by_country[country][(model, day)]
                    value = calculate_metric_value(bucket, metric_key, test_drive_expression)
                    row.append(value or None)
                rows.append(tuple(row))
    return rows, headers


def build_country_rows(
    records,
    model_equivalences,
    start_date,
    end_date,
    report_models=None,
    test_drive_expression=None,
):
    dates = set()
    counts = defaultdict(Counter)
    old_dates = set()
    old_counts = defaultdict(Counter)
    unknown_request_types = Counter()
    rows_in_period = 0
    old_rows_in_period = 0
    for record in records:
        create_date = record["create_date"]
        if create_date < start_date or create_date > end_date:
            continue

        rows_in_period += 1
        model = model_equivalences.get(record["id"], "UNMAPPED")
        if text_key(model) == "OLD":
            old_rows_in_period += 1
            old_dates.add(create_date)
            bucket = old_counts[(model, create_date)]
        else:
            dates.add(create_date)
            bucket = counts[(model, create_date)]
        request_type = record["request_type"]
        if request_type == "OFFER REQUEST":
            bucket["offer_request"] += 1
        elif request_type == "TEST DRIVE REQUEST":
            bucket["test_drive_request"] += 1
        elif request_type == "DEALER LEADS":
            bucket["dealer_leads"] += 1
        else:
            unknown_request_types[request_type] += 1

    rows, headers = build_metric_rows(
        counts,
        dates,
        report_models,
        test_drive_expression,
    )
    return (
        rows,
        headers,
        old_counts,
        old_dates,
        rows_in_period,
        old_rows_in_period,
        unknown_request_types,
    )


def build_unmapped_rows(lead_sources, model_equivalences):
    rows = []
    for country, source in lead_sources.items():
        for lead_id, occurrences in source["id_counts"].items():
            if lead_id not in model_equivalences:
                rows.append((country, lead_id, occurrences))
    return sorted(rows, key=lambda row: (row[0], row[1]))


def write_output(country_outputs, old_output, unmapped_rows):
    report_progress("Writing Leads analysis workbook...")
    workbook = Workbook(write_only=True)
    try:
        for country in ("SP", "PT"):
            rows, headers = country_outputs[country]
            worksheet = workbook.create_sheet(f"LEADS_{country}")
            widths = calculate_column_widths(headers, rows, list, format_value)
            prepare_worksheet(worksheet, headers, widths, 36)
            for row in rows:
                append_row(worksheet, row, SHORT_DATE_FORMAT)

        old_rows, old_headers = old_output
        worksheet = workbook.create_sheet("OLD_SP OLD_PT")
        widths = calculate_column_widths(old_headers, old_rows, list, format_value)
        prepare_worksheet(worksheet, old_headers, widths, 36)
        for row in old_rows:
            append_row(worksheet, row, SHORT_DATE_FORMAT)

        unmapped_headers = ["country", "id", "occurrences"]
        worksheet = workbook.create_sheet("UNMAPPED_MODELS")
        widths = calculate_column_widths(unmapped_headers, unmapped_rows, list, format_value)
        prepare_worksheet(worksheet, unmapped_headers, widths, 44)
        for row in unmapped_rows:
            append_row(worksheet, row, SHORT_DATE_FORMAT)

        save_workbook_atomically(workbook, OUTPUT_XLSX_PATH)
    finally:
        workbook.close()
    report_progress(f"Output workbook written: {OUTPUT_XLSX_PATH}")


def main(start_date, end_date, test_drive_formula=DEFAULT_TEST_DRIVE_FORMULA):
    start_date = parse_period_date(start_date)
    end_date = parse_period_date(end_date)
    if start_date > end_date:
        raise ValueError("Leads start date cannot be after the end date.")
    test_drive_formula, test_drive_expression = parse_test_drive_formula(test_drive_formula)

    report_progress("Starting Leads analysis...")
    report_progress(f"Leads period: {start_date.isoformat()} to {end_date.isoformat()}")
    report_progress(f"Test drive formula: {test_drive_formula}")
    lead_sources = {country: load_leads(country) for country in ("SP", "PT")}
    model_equivalences, duplicate_model_ids = load_model_equivalences()
    report_models = build_reporting_models(model_equivalences)

    country_outputs = {}
    old_counts_by_country = {}
    old_dates_by_country = {}
    unknown_request_types = {}
    for country, source in lead_sources.items():
        (
            rows,
            headers,
            old_counts,
            old_dates,
            rows_in_period,
            old_rows_in_period,
            unknown_types,
        ) = build_country_rows(
            source["records"],
            model_equivalences,
            start_date,
            end_date,
            report_models,
            test_drive_expression,
        )
        country_outputs[country] = (rows, headers)
        old_counts_by_country[country] = old_counts
        old_dates_by_country[country] = old_dates
        source["rows_in_period"] = rows_in_period
        source["old_rows_in_period"] = old_rows_in_period
        unknown_request_types[country] = unknown_types

    unmapped_rows = build_unmapped_rows(lead_sources, model_equivalences)
    for country, source in lead_sources.items():
        if source["invalid_dates"]:
            print(f"WARNING: Leads {country} rows with invalid create date: {len(source['invalid_dates']):,}")
            print("; ".join(source["invalid_dates"][:20]))
        if source["skipped_without_id"]:
            print(f"WARNING: Leads {country} rows without an interest model: {source['skipped_without_id']:,}")
        if unknown_request_types[country]:
            print(f"WARNING: Leads {country} unknown request types: {sum(unknown_request_types[country].values()):,}")
            for request_type, count in sorted(unknown_request_types[country].items()):
                print(f"{request_type}; {count:,}")
    if unmapped_rows:
        print(f"WARNING: Leads IDs without model equivalence: {len(unmapped_rows):,}")
        for country, lead_id, occurrences in unmapped_rows[:50]:
            print(f"{country}; {lead_id}; {occurrences:,}")

    old_output = build_old_rows(
        old_counts_by_country,
        old_dates_by_country,
        test_drive_expression,
    )
    write_output(country_outputs, old_output, unmapped_rows)
    print("PROCESS_SUMMARY")
    print(f"period_start: {start_date.isoformat()}")
    print(f"period_end: {end_date.isoformat()}")
    print(f"model_equivalences: {len(model_equivalences):,}")
    print(f"test_drive_formula: {test_drive_formula}")
    print(f"duplicate_model_equivalence_rows: {duplicate_model_ids:,}")
    for country in ("SP", "PT"):
        source = lead_sources[country]
        print(f"leads_{country.lower()}_rows_scanned: {source['rows_scanned']:,}")
        print(f"leads_{country.lower()}_rows_in_period: {source['rows_in_period']:,}")
        print(f"leads_{country.lower()}_old_rows_in_period: {source['old_rows_in_period']:,}")
        print(f"leads_{country.lower()}_invalid_dates: {len(source['invalid_dates']):,}")
    print(f"unmapped_model_ids: {len(unmapped_rows):,}")
    print(f"OUTPUT_XLSX: {OUTPUT_XLSX_PATH}")


if __name__ == "__main__":
    main(date.today(), date.today())

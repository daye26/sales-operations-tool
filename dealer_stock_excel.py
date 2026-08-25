from collections import Counter, defaultdict
import csv
from datetime import date, datetime, time
from pathlib import Path
import re
import sys
import unicodedata
import warnings

from openpyxl import Workbook, load_workbook

from excel_sheet_selection import select_active_then_sheet1
from excel_output import append_row, calculate_column_widths, prepare_worksheet, save_workbook_atomically
import free_cars_history
from port_resolution import resolve_port
import tabular_normalization as tabular
import vehicle_tracking_cache as tracking_cache
import vehicle_tracking_loader


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

BASE_EXCEL_DIR = Path.cwd()
LOGISTICS_DB_PATH = BASE_EXCEL_DIR / "BASE DE DATOS LOGISTICA.xlsx"
OUTPUT_XLSX_PATH = BASE_EXCEL_DIR / "dealer_stock_result.xlsx"
PROGRESS_CALLBACK = None
SHORT_DATE_FORMAT = "yyyy-mm-dd"
CACHE_DIR = tracking_cache.default_cache_dir()
VEHICLE_TRACKING_CACHE_PATH = CACHE_DIR / "vehicle_tracking.pkl"
VEHICLE_TRACKING_CACHE_VERSION = tracking_cache.CACHE_VERSION

EXCEL_PATHS = {
    "vehicle_tracking": BASE_EXCEL_DIR / "VehicleTracking.xlsx",
    "mc_norm": BASE_EXCEL_DIR / "material code.xlsx",
    "dealer_info": BASE_EXCEL_DIR / "Dealer Info.xlsx",
    "not_allocated": BASE_EXCEL_DIR / "Cars not allocated.xlsx",
    "newport": BASE_EXCEL_DIR / "NEWport.xlsx",
    "orders": BASE_EXCEL_DIR / "Dealer Orders.xlsx",
    "registration": BASE_EXCEL_DIR / "ANT.xlsx",
    "unavailable": BASE_EXCEL_DIR / "unavailable.xlsx",
    "derogation": BASE_EXCEL_DIR / "derogation_list.xlsx",
    "reservation": BASE_EXCEL_DIR / "Vehicle_Reservation.xlsx",
    "in_port": LOGISTICS_DB_PATH,
    "already_gate_out": LOGISTICS_DB_PATH,
}

SHEET_NAMES = {
    "not_allocated": "Sheet1",
    "orders": "Dealer Orders",
    "registration": "LISTADO",
    "derogation": "derogation list",
    "in_port": "Port Stock",
    "already_gate_out": "Already Gate Out Bis",
}

OLD_DEROGATION_CUTOFF = None
EXCLUDED_BACK_ORDER_DEALER = "Internal Dealer"

DEALER_STOCK_COLUMNS = [
    "vin",
    "material_code",
    "description",
    "p_series",
    "p_trim",
    "int_color",
    "ext_color",
    "port",
    "vessel_name",
    "eta",
    "group_name",
    "dsn",
    "status",
    "dealer",
    "or_number",
    "acc_type",
    "owner_name",
    "client",
    "or_date",
    "crm",
    "o_match",
    "supplier",
    "gate_out",
    "invoice_date",
    "gate_out_days",
    "duration_logic",
    "derogation",
    "production_date",
    "prod_days",
]

HEADER_ALIASES = {
    "vin": ["vin", "vin #"],
    "material_code": ["material code", "vehicle material code"],
    "description": ["description(local)", "description"],
    "port": ["current warehouse", "port"],
    "vessel": ["name of vessel", "vessel"],
    "eta": ["purchase eta", "eta"],
    "dsn": ["delivery short name", "dsn"],
    "status": ["current status", "status"],
    "gate_in": ["gate in date", "gate in", "gate_in"],
    "gate_out": ["gate out date", "actual gate out"],
    "production_date": ["production date"],
    "sap": ["customer# (sap code)", "customer#", "sap"],
    "invoice_date": ["sales invoice date"],
    "address": ["detail delivery address", "address"],
    "city": ["city"],
    "code": ["material code", "code"],
    "p_series": ["product series", "model"],
    "p_trim": ["product model", "trim"],
    "int_color": ["interial color", "interior color"],
    "ext_color": ["exterial color", "exterior color"],
    "dealer_code": ["dealer code"],
    "dealer_info_code": ["code"],
    "dealer": ["sales dealer: account name", "dealer"],
    "or_number": ["order number"],
    "acc_type": ["account record type"],
    "owner_name": ["owner name"],
    "client": ["account name: account name"],
    "or_date": ["activated date"],
    "vehicle_vin": ["vehicle vin"],
    "country": ["country"],
    "customer_country": ["customer country"],
    "tag": ["tag name", "tag"],
    "related_order": ["related order#", "related order", "related_order"],
    "reserved_so": ["reserved so #", "reserved so", "reserved_so"],
    "dn_create_time": ["dn create time", "dn_create_time"],
    "allocation_date": ["allocation date", "allocate date", "allocation_date"],
    "group_name": ["group_name", "group name"],
    "submit_date": ["submit date"],
    "m_date": ["m_date", "registration date"],
    "note": ["note"],
    "supplier": ["supplier"],
}


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


is_missing = tabular.is_missing


def report_progress(message):
    if PROGRESS_CALLBACK is None:
        print(message)
        return

    try:
        PROGRESS_CALLBACK(message)
    except Exception:
        pass


def vehicle_tracking_file_signature():
    return tracking_cache.vehicle_tracking_file_signature(EXCEL_PATHS["vehicle_tracking"])


def clear_vehicle_tracking_cache():
    return tracking_cache.clear_cache(VEHICLE_TRACKING_CACHE_PATH)


def load_vehicle_tracking_cache(signature):
    return tracking_cache.load_cache(VEHICLE_TRACKING_CACHE_PATH, signature, report_progress)


def write_vehicle_tracking_cache(signature, by_vin):
    tracking_cache.write_cache(
        VEHICLE_TRACKING_CACHE_PATH,
        signature,
        by_vin,
        report_progress,
        vehicle_tracking_loader.ALL_SOURCE_FIELDS,
    )


format_value = tabular.format_value
normalize_header = tabular.normalize_header
text_key = tabular.text_key
code_key = tabular.code_key
vin_key = tabular.vin_key


def header_index(columns, column_name, required=True):
    return tabular.header_index(columns, HEADER_ALIASES, column_name, required)


def build_indexes(columns, required_columns, optional_columns=None):
    return tabular.build_indexes(columns, HEADER_ALIASES, required_columns, optional_columns)


row_value = tabular.row_value
max_required_col = tabular.max_required_col


def open_sheet(key):
    path = EXCEL_PATHS[key]
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    selected_sheet = SHEET_NAMES.get(key)
    if key == "not_allocated":
        worksheet = select_active_then_sheet1(workbook, HEADER_ALIASES, ("vin", "note"))
    else:
        worksheet = workbook[selected_sheet] if selected_sheet in workbook.sheetnames else workbook.active
    worksheet.reset_dimensions()
    return workbook, worksheet


def read_header(worksheet):
    return [format_value(value) for value in next(worksheet.iter_rows(values_only=True))]


def sniff_dialect(path):
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as file:
        sample = file.read(4096)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def iter_csv_dicts(path):
    dialect = sniff_dialect(path)
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as file:
        reader = csv.reader(file, dialect)
        headers = next(reader)
        indexes = {normalize_header(header): index for index, header in enumerate(headers)}
        for row in reader:
            if not any(not is_missing(value) for value in row):
                continue
            yield headers, indexes, row


def iter_excel_sheet_dicts(path, sheet_name=None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        worksheet.reset_dimensions()
        headers = [format_value(value) for value in next(worksheet.iter_rows(values_only=True))]
        indexes = {normalize_header(header): index for index, header in enumerate(headers)}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            yield headers, indexes, row
    finally:
        workbook.close()


def iter_table_rows(key):
    path = EXCEL_PATHS[key]
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        yield from iter_excel_sheet_dicts(path, SHEET_NAMES.get(key))
    else:
        yield from iter_csv_dicts(path)


def csv_value(row, indexes, *aliases):
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    for alias in normalized_aliases:
        index = indexes.get(alias)
        if index is not None and index < len(row):
            return row[index]
    return None


def to_datetime(value):
    if is_missing(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = format_value(value)
    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H/%M",
        "%d/%m/%Y,%H:%M",
        "%d/%m/%Y,%H/%M",
        "%d/%m/%Y, %H:%M",
        "%d/%m/%Y, %H:%M:%S",
        "%d/%m/%Y, %H/%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass
    return datetime.fromisoformat(text)


def to_date(value):
    parsed = to_datetime(value)
    if parsed is None:
        return None
    return parsed.date()


def days_since(value):
    parsed = to_date(value)
    if parsed is None:
        return None
    return (date.today() - parsed).days


def date_sort_key(value):
    parsed = to_datetime(value)
    return parsed if parsed is not None else datetime.max


def material_aggr(mc):
    return text_key(
        (mc.get("aggr") or "")
        or (
            mc.get("p_series", "")
            + mc.get("p_trim", "")
            + mc.get("int_color", "")
            + mc.get("ext_color", "")
        )
    )


def dealer_stock_aggr(row):
    return text_key(
        row.get("group_name", "")
        + row.get("p_series", "")
        + row.get("p_trim", "")
        + row.get("int_color", "")
        + row.get("ext_color", "")
    )


def load_mc_norm():
    workbook, worksheet = open_sheet("mc_norm")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            ["code", "p_series", "p_trim", "int_color", "ext_color"],
            optional_columns=["description"],
        )
        by_code = {}
        counter = Counter()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            code = code_key(row_value(row, indexes, "code"))
            if is_missing(code):
                continue
            counter[code] += 1
            by_code[code] = {
                "code": code,
                "p_series": format_value(row_value(row, indexes, "p_series")),
                "p_trim": format_value(row_value(row, indexes, "p_trim")),
                "int_color": format_value(row_value(row, indexes, "int_color")),
                "ext_color": format_value(row_value(row, indexes, "ext_color")),
                "aggr": "",
            }

        duplicated = [code for code, count in counter.items() if count > 1]
        if duplicated:
            raise ValueError(f"mc_norm duplicated material codes: {', '.join(sorted(duplicated)[:20])}")
        return by_code
    finally:
        workbook.close()


def load_dealer_info():
    workbook, worksheet = open_sheet("dealer_info")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["dsn", "dealer_info_code", "sap", "dealer", "group_name", "country"])
        by_sap = defaultdict(list)
        sp_saps = set()
        by_sap_seen = defaultdict(set)
        by_code = {}

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            dealer = {
                "dsn": format_value(row_value(row, indexes, "dsn")),
                "code": code_key(row_value(row, indexes, "dealer_info_code")),
                "sap": code_key(row_value(row, indexes, "sap")),
                "dealer": format_value(row_value(row, indexes, "dealer")),
                "group_name": format_value(row_value(row, indexes, "group_name")),
                "country": code_key(row_value(row, indexes, "country")),
            }
            if dealer["sap"]:
                if dealer["country"] == "SP":
                    sp_saps.add(dealer["sap"])
                business_key = (dealer["group_name"], dealer["dsn"], dealer["country"])
                if business_key not in by_sap_seen[dealer["sap"]]:
                    by_sap_seen[dealer["sap"]].add(business_key)
                    by_sap[dealer["sap"]].append(dealer)
            if dealer["code"]:
                by_code[dealer["code"]] = dealer

        return by_sap, sp_saps, by_code
    finally:
        workbook.close()


def load_vin_set_from_table(key, vin_aliases, optional=False):
    path = EXCEL_PATHS[key]
    if optional and not path.exists():
        print(f"WARNING: optional file not found, skipped: {path}")
        return set()
    if not path.exists():
        raise FileNotFoundError(path)

    vins = set()
    for _, indexes, row in iter_table_rows(key):
        vin = vin_key(csv_value(row, indexes, *vin_aliases))
        if vin:
            vins.add(vin)
    return vins


def load_registration_dates():
    registration_dates = {}
    for _, indexes, row in iter_table_rows("registration"):
        vin = vin_key(csv_value(row, indexes, "vin"))
        if not vin:
            continue
        registration_dates.setdefault(vin, csv_value(row, indexes, "m_date"))
    return registration_dates


def load_not_allocated_notes():
    workbook, worksheet = open_sheet("not_allocated")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["vin", "note"])
        notes = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            vin = vin_key(row_value(row, indexes, "vin"))
            if vin:
                notes[vin] = format_value(row_value(row, indexes, "note"))
        return notes
    finally:
        workbook.close()


def load_not_allocated_ports():
    workbook, worksheet = open_sheet("not_allocated")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["vin"], optional_columns=["port"])
        ports_by_vin = {}
        conflicting_vins = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row_value(row, indexes, "vin"))
            port = format_value(row_value(row, indexes, "port"))
            if not vin or not port:
                continue
            previous_port = ports_by_vin.get(vin)
            if previous_port and text_key(previous_port) != text_key(port):
                conflicting_vins.add(vin)
                continue
            ports_by_vin.setdefault(vin, port)
    finally:
        workbook.close()

    if conflicting_vins:
        report_progress(
            "WARNING: Cars not allocated VINs with conflicting ports: "
            + "; ".join(sorted(conflicting_vins)[:20])
        )
    return ports_by_vin


def load_newport_ports():
    path = EXCEL_PATHS.get("newport")
    if path is None or not Path(path).exists():
        report_progress("WARNING: NEWport file not found; using the remaining port sources")
        return {}
    workbook, worksheet = open_sheet("newport")
    try:
        columns = read_header(worksheet)
        try:
            indexes = build_indexes(columns, ["vin", "port"])
        except ValueError:
            report_progress(
                "WARNING: NEWport does not contain VIN and port columns; "
                "using the remaining port sources"
            )
            return {}
        ports_by_vin = {}
        duplicate_vins = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row_value(row, indexes, "vin"))
            port = format_value(row_value(row, indexes, "port"))
            if not vin or not port:
                raise ValueError("NEWport has a row with an empty VIN or port")
            if vin in ports_by_vin:
                duplicate_vins.add(vin)
            ports_by_vin[vin] = port
    finally:
        workbook.close()

    if duplicate_vins:
        raise ValueError("NEWport has duplicated VINs: " + ", ".join(sorted(duplicate_vins)[:20]))
    return ports_by_vin


def load_derogation_rows():
    rows = []
    for _, indexes, row in iter_table_rows("derogation"):
        vin = vin_key(csv_value(row, indexes, "VIN"))
        if not vin:
            continue
        rows.append(
            {
                "vin": vin,
                "submit_date": csv_value(row, indexes, "Submit Date"),
            }
        )
    return rows


def load_reservation_vins():
    return load_vin_set_from_table("reservation", ["vin", "vin #", "vin#"], optional=True)


def load_vehicle_tracking():
    return vehicle_tracking_loader.load_vehicle_tracking(
        EXCEL_PATHS["vehicle_tracking"],
        VEHICLE_TRACKING_CACHE_PATH,
        report_progress,
        required_fields=(
            "vin",
            "material_code",
            "description",
            "port",
            "vessel",
            "eta",
            "dsn",
            "sap",
            "status",
            "gate_in",
            "gate_out",
            "production_date",
            "invoice_date",
            "customer_country",
        ),
    )


def load_vehicle_tracking_lookup_for_vins(vins):
    lookup = {}
    for vin, row in load_vehicle_tracking().items():
        if vin not in vins:
            continue
        lookup[vin] = {
            "material_code": code_key(row.get("material_code")),
            "dsn": format_value(row.get("dsn")),
            "country": format_value(row.get("country")),
        }

    report_progress(f"vehicle_tracking_derogation_rows_loaded: {len(lookup):,}")
    return lookup


def format_date_text(value):
    parsed = to_date(value)
    if parsed is not None:
        return parsed
    return format_value(value)


def build_derogation_list_rows(
    mc_norm,
    derogation_rows,
    vehicle_tracking_by_vin,
    reservation_vins,
    not_allocated_notes,
    registration_dates,
):
    rows = []
    seen = set()
    for derogation in derogation_rows:
        vin = derogation["vin"]
        tracking = vehicle_tracking_by_vin.get(vin, {})
        material_code = tracking.get("material_code", "")
        mc = mc_norm.get(material_code, {})
        row = (
            vin,
            "Y",
            mc.get("code", material_code),
            format_date_text(derogation["submit_date"]),
            format_date_text(registration_dates.get(vin)),
            mc.get("p_series", ""),
            mc.get("p_trim", ""),
            tracking.get("dsn", ""),
            tracking.get("country", ""),
            "Y" if vin in reservation_vins else "",
            not_allocated_notes.get(vin, ""),
        )
        if row in seen:
            continue
        seen.add(row)
        rows.append(row)

    return sorted(rows, key=lambda row: row[0])


def make_stock_row(vt, mc, dealer):
    production_date = to_date(vt["production_date"])
    invoice_date = to_date(vt["invoice_date"])
    gate_out = to_date(vt["gate_out"])
    prod_days = days_since(production_date)
    gate_out_days = days_since(invoice_date)
    duration_logic = ""
    if gate_out_days is not None:
        if gate_out_days > 360:
            duration_logic = ">360 days"
        elif gate_out_days > 300:
            duration_logic = ">300 days"
        elif gate_out_days > 270:
            duration_logic = ">270 days"
        elif gate_out_days > 180:
            duration_logic = ">180 days"
        else:
            duration_logic = "<180 days"

    return {
        "vin": vt["vin"],
        "material_code": vt["material_code"],
        "description": vt["description"],
        "p_series": mc["p_series"],
        "p_trim": mc["p_trim"],
        "int_color": mc["int_color"],
        "ext_color": mc["ext_color"],
        "port": vt["port"],
        "vessel_name": vt["vessel_name"],
        "eta": to_datetime(vt["eta"]),
        "group_name": dealer["group_name"],
        "dsn": vt["dsn"],
        "status": vt["status"],
        "dealer": "",
        "or_number": "",
        "acc_type": "",
        "owner_name": "",
        "client": "",
        "or_date": None,
        "crm": "",
        "o_match": "",
        "supplier": "",
        "gate_out": gate_out,
        "invoice_date": invoice_date,
        "gate_out_days": gate_out_days,
        "duration_logic": duration_logic,
        "derogation": "",
        "production_date": production_date,
        "prod_days": prod_days,
    }


def stock_tuple(row):
    return tuple(row.get(column) for column in DEALER_STOCK_COLUMNS)


def load_dealer_stock_pool(
    mc_norm,
    dealer_info_by_sap,
    sp_saps,
    registration_vins,
    unavailable_vins,
    derogation_vins=None,
    port_stock_ports=None,
    newport_ports=None,
    not_allocated_ports=None,
):
    derogation_vins = derogation_vins or set()
    port_stock_ports = port_stock_ports or {}
    newport_ports = newport_ports or {}
    not_allocated_ports = not_allocated_ports or {}
    vehicle_tracking = load_vehicle_tracking()
    stock_rows = []
    derogation_vehicle_tracking = {}
    unique_rows = set()
    missing_mc_norm = []
    missing_group_name = []
    scanned = 0

    report_progress("Processing VehicleTracking rows for dealer stock...")
    for scanned, vt in enumerate(vehicle_tracking.values(), start=1):
        if scanned % 25000 == 0:
            report_progress(f"vehicle_tracking_rows_processed: {scanned}")

        vin = vin_key(vt.get("vin"))
        material_code = code_key(vt.get("material_code"))
        dsn = format_value(vt.get("dsn"))
        sap = code_key(vt.get("sap"))
        status = format_value(vt.get("status"))
        if vin in derogation_vins:
            derogation_vehicle_tracking[vin] = {
                "material_code": material_code,
                "dsn": dsn,
                "country": format_value(vt.get("country")),
            }
        if not vin:
            continue
        if vin in registration_vins or vin in unavailable_vins:
            continue
        # Equivalent to: EXISTS (dealer_info d WHERE v.sap = d.sap AND d.country = 'SP').
        if not sap or sap not in sp_saps:
            continue
        if is_missing(status) or status.upper() == "OFFLINE":
            continue

        mc = mc_norm.get(material_code)
        if mc is None or any(is_missing(mc[column]) for column in ["p_series", "p_trim", "int_color", "ext_color"]):
            missing_mc_norm.append(f"{vin};{material_code}")
            continue

        stock_vt = {
            "vin": vin,
            "material_code": material_code,
            "description": format_value(vt.get("description")),
            "port": resolve_port(
                port_stock_ports.get(vin),
                newport_ports.get(vin),
                not_allocated_ports.get(vin),
                vt.get("port"),
            ),
            "vessel_name": format_value(vt.get("vessel_name")),
            "eta": vt.get("eta"),
            "dsn": dsn,
            "status": status,
            "gate_out": vt.get("gate_out"),
            "production_date": vt.get("production_date"),
            "invoice_date": vt.get("invoice_date"),
        }
        for dealer in dealer_info_by_sap[sap]:
            if is_missing(dealer["group_name"]):
                missing_group_name.append(f"{vin};{dsn}")
                continue
            stock_row = make_stock_row(stock_vt, mc, dealer)
            row_key = stock_tuple(stock_row)
            if row_key in unique_rows:
                continue
            unique_rows.add(row_key)
            stock_rows.append(stock_row)

    if missing_mc_norm:
        detail = "\n".join(missing_mc_norm[:50])
        raise ValueError(f"dealer_stock_pool missing mc_norm rows: {len(missing_mc_norm)}\n{detail}")

    if missing_group_name:
        detail = "\n".join(sorted(set(missing_group_name))[:50])
        raise ValueError(
            f"dealer_stock_pool has VINs without dealer_info.group_name: {len(set(missing_group_name))}\n"
            f"{detail}"
        )

    duplicate_vins = [
        vin for vin, count in Counter(row["vin"] for row in stock_rows).items() if count > 1
    ]
    if duplicate_vins:
        detail = "\n".join(sorted(duplicate_vins)[:50])
        raise ValueError(f"dealer_stock_pool duplicated VINs: {len(duplicate_vins)}\n{detail}")

    report_progress(f"vehicle_tracking_rows_processed: {scanned}")
    return stock_rows, derogation_vehicle_tracking


def load_orders(mc_norm, dealer_info_by_code):
    path = EXCEL_PATHS["orders"]
    orders = []
    missing_mc_norm = []
    missing_dealer_info_without_vin = []
    for _, indexes, row in iter_table_rows("orders"):
        material_code = code_key(csv_value(row, indexes, "Material Code"))
        dealer_code = code_key(csv_value(row, indexes, "Dealer Code"))
        order_number = format_value(csv_value(row, indexes, "Order Number"))
        sales_dealer = format_value(csv_value(row, indexes, "Sales Dealer: Account Name"))
        vehicle_vin = vin_key(csv_value(row, indexes, "Vehicle VIN"))
        mc = mc_norm.get(material_code)
        dealer_info = dealer_info_by_code.get(dealer_code)
        if mc is None:
            missing_mc_norm.append(f"{order_number};{material_code}")
            continue
        if dealer_info is None and not vehicle_vin:
            missing_dealer_info_without_vin.append(f"{order_number};{dealer_code};{sales_dealer}")
        orders.append(
            {
                "d_code": dealer_code,
                "dealer": sales_dealer,
                "or_number": order_number,
                "acc_type": format_value(csv_value(row, indexes, "Account Record Type")),
                "owner_name": format_value(csv_value(row, indexes, "Owner Name")),
                "client": format_value(csv_value(row, indexes, "Account Name: Account Name")),
                "or_date": to_datetime(csv_value(row, indexes, "Activated Date")),
                "material_code": material_code,
                "vin": vehicle_vin,
                "p_trim": format_value(csv_value(row, indexes, "Product Model")),
                "mc": mc,
                "dealer_info": dealer_info,
            }
        )

    if missing_mc_norm:
        detail = "\n".join(missing_mc_norm[:50])
        raise ValueError(f"orders missing mc_norm rows: {len(missing_mc_norm)}\n{detail}")
    if missing_dealer_info_without_vin:
        print(
            "SOFT WARNING: orders without VIN missing dealer_info rows "
            f"(ignored for back-order matching): {len(missing_dealer_info_without_vin)}"
        )
        for item in missing_dealer_info_without_vin[:20]:
            print(item)
    return orders


def apply_direct_order_matches(stock_rows, orders):
    direct_order_vin_counts = Counter(order["vin"] for order in orders if order["vin"])
    duplicated_direct_order_vins = sorted(
        vin for vin, count in direct_order_vin_counts.items() if count > 1
    )
    if duplicated_direct_order_vins:
        details = []
        for vin in duplicated_direct_order_vins[:20]:
            order_numbers = [
                format_value(order["or_number"])
                for order in orders
                if order["vin"] == vin
            ]
            details.append(f"{vin}: {', '.join(order_numbers)}")
        raise ValueError(
            "orders has duplicated direct Vehicle VIN values:\n"
            + "\n".join(details)
        )

    by_vin = {row["vin"]: row for row in stock_rows}
    for order in orders:
        if not order["vin"] or order["vin"] not in by_vin:
            continue
        row = by_vin[order["vin"]]
        row.update(
            {
                "dealer": order["dealer"],
                "or_number": order["or_number"],
                "acc_type": order["acc_type"],
                "owner_name": order["owner_name"],
                "client": order["client"],
                "or_date": order["or_date"],
                "crm": "Y",
                "o_match": "Y",
            }
        )


def apply_back_order_matches(stock_rows, orders):
    stock_by_key = defaultdict(list)
    for row in stock_rows:
        if row["or_number"]:
            continue
        stock_by_key[dealer_stock_aggr(row)].append(row)
    for rows in stock_by_key.values():
        rows.sort(key=lambda row: (date_sort_key(row["eta"]), row["vin"]))

    order_by_key = defaultdict(list)
    min_date = datetime(2025, 1, 1)
    for order in orders:
        if order["vin"]:
            continue
        if text_key(order["dealer"]) == text_key(EXCLUDED_BACK_ORDER_DEALER):
            continue
        if order["dealer_info"] is None:
            continue
        if order["or_date"] is None or order["or_date"] < min_date:
            continue
        dealer_info = order["dealer_info"] or {}
        key = text_key((dealer_info.get("group_name") or "") + material_aggr(order["mc"]))
        order_by_key[key].append(order)
    for rows in order_by_key.values():
        rows.sort(key=lambda order: (date_sort_key(order["or_date"]), order["or_number"]))

    for key, stock_candidates in stock_by_key.items():
        for row, order in zip(stock_candidates, order_by_key.get(key, [])):
            row.update(
                {
                    "dealer": order["dealer"],
                    "or_number": order["or_number"],
                    "acc_type": order["acc_type"],
                    "owner_name": order["owner_name"],
                    "client": order["client"],
                    "or_date": order["or_date"],
                    "crm": "N",
                    "o_match": "Y",
                }
            )

    for row in stock_rows:
        if not row["o_match"]:
            row["o_match"] = "N"


def apply_derogations(stock_rows, registration_vins, derogation_rows=None):
    if derogation_rows is None:
        path = EXCEL_PATHS["derogation"]
        if not path.exists():
            print(f"WARNING: file not found, skipped: {path}")
            return
        derogation_rows = load_derogation_rows()

    by_vin = {row["vin"]: row for row in stock_rows}
    for derogation in derogation_rows:
        vin = derogation["vin"]
        if not vin or vin in registration_vins or vin not in by_vin:
            continue
        submit_date = derogation["submit_date"]
        if not is_missing(submit_date):
            by_vin[vin]["derogation"] = to_date(submit_date) or format_value(submit_date)


def supplier_status(supplier):
    supplier_text = format_value(supplier)
    if supplier_text.upper() in {
        "",
        "FINDING",
        "CANCELLED",
        "WAITING FOR ASSIGMENT",
        "WAITING FOR ASSIGNMENT",
    }:
        return "", "Without supplier"
    return supplier_text, "With supplier"


def apply_port_stock_row(by_vin, indexes, csv_row, only_status=None, skip_status=None):
    vin = vin_key(csv_value(csv_row, indexes, "VIN"))
    if vin not in by_vin:
        return
    row = by_vin[vin]
    if only_status is not None and row["status"] != only_status:
        return
    if skip_status is not None and row["status"] == skip_status:
        return
    supplier, status = supplier_status(csv_value(csv_row, indexes, "Supplier"))
    row["supplier"] = supplier
    row["status"] = status


def apply_port_stock(stock_rows, key, only_status=None, skip_status=None):
    path = EXCEL_PATHS[key]
    if not path.exists():
        print(f"WARNING: file not found, skipped: {path}")
        return

    by_vin = {row["vin"]: row for row in stock_rows}
    for _, indexes, csv_row in iter_table_rows(key):
        apply_port_stock_row(by_vin, indexes, csv_row, only_status, skip_status)


def apply_port_stock_sources(stock_rows):
    in_port_path = EXCEL_PATHS["in_port"]
    gate_out_path = EXCEL_PATHS["already_gate_out"]
    shared_excel_file = (
        in_port_path.exists()
        and gate_out_path.exists()
        and in_port_path.resolve() == gate_out_path.resolve()
        and in_port_path.suffix.lower() in {".xlsx", ".xlsm"}
    )
    if not shared_excel_file:
        apply_port_stock(stock_rows, "in_port", skip_status="Shipping")
        apply_port_stock(stock_rows, "already_gate_out", only_status="Booked")
        return

    by_vin = {row["vin"]: row for row in stock_rows}
    workbook = load_workbook(in_port_path, read_only=True, data_only=True)
    try:
        for key, only_status, skip_status in (
            ("in_port", None, "Shipping"),
            ("already_gate_out", "Booked", None),
        ):
            worksheet = workbook[SHEET_NAMES[key]]
            worksheet.reset_dimensions()
            headers = [format_value(value) for value in next(worksheet.iter_rows(values_only=True))]
            indexes = {normalize_header(header): index for index, header in enumerate(headers)}
            for csv_row in worksheet.iter_rows(min_row=2, values_only=True):
                if not any(not is_missing(value) for value in csv_row):
                    continue
                apply_port_stock_row(by_vin, indexes, csv_row, only_status, skip_status)
    finally:
        workbook.close()


def excel_value(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def write_sheet(workbook, title, columns, rows):
    worksheet = workbook.create_sheet(title)
    def row_values(row):
        if isinstance(row, dict):
            return [excel_value(row.get(column)) for column in columns]
        return [excel_value(value) for value in row]

    column_widths = calculate_column_widths(columns, rows, row_values, format_value)
    prepare_worksheet(worksheet, columns, column_widths, 45)
    for row in rows:
        append_row(worksheet, row_values(row), SHORT_DATE_FORMAT)


def write_output(stock_rows, derogation_list_rows):
    workbook = Workbook(write_only=True)
    stock_rows = sorted(stock_rows, key=lambda row: row["vin"])
    old_derogations = []
    for row in stock_rows:
        derogation_date = to_date(row["derogation"])
        if (
            OLD_DEROGATION_CUTOFF is not None
            and derogation_date is not None
            and derogation_date < OLD_DEROGATION_CUTOFF
        ):
            old_derogations.append((row["vin"], row["p_series"], row["dsn"]))

    try:
        write_sheet(workbook, "DEALER_STOCK", DEALER_STOCK_COLUMNS, stock_rows)
        write_sheet(workbook, "OLD_DEROGATIONS", ["vin", "p_series", "dsn"], old_derogations)
        write_sheet(
            workbook,
            "DEROGATION_LIST",
            [
                "vin",
                "Y",
                "material_code",
                "submit_date",
                "registration_date",
                "p_series",
                "p_trim",
                "dsn",
                "country",
                "reservation",
                "note",
            ],
            derogation_list_rows,
        )
        save_workbook_atomically(workbook, OUTPUT_XLSX_PATH)
    finally:
        workbook.close()

    print("PROCESS_SUMMARY")
    print(f"dealer_stock_rows: {len(stock_rows)}")
    print(f"duplicated_vins: {sum(1 for count in Counter(row['vin'] for row in stock_rows).values() if count > 1)}")
    print(f"crm_y: {sum(1 for row in stock_rows if row['crm'] == 'Y')}")
    print(f"crm_n: {sum(1 for row in stock_rows if row['crm'] == 'N')}")
    print(f"o_match_y: {sum(1 for row in stock_rows if row['o_match'] == 'Y')}")
    print(f"o_match_n: {sum(1 for row in stock_rows if row['o_match'] == 'N')}")
    cutoff_text = (
        format_date_text(OLD_DEROGATION_CUTOFF)
        if OLD_DEROGATION_CUTOFF is not None
        else "not set"
    )
    print(f"old_derogation_cutoff: {cutoff_text}")
    print(f"old_derogations: {len(old_derogations)}")
    print(f"derogation_list_rows: {len(derogation_list_rows)}")
    print(f"with_supplier: {sum(1 for row in stock_rows if row['status'] == 'With supplier')}")
    print(f"without_supplier: {sum(1 for row in stock_rows if row['status'] == 'Without supplier')}")
    print(f"OUTPUT_XLSX: {OUTPUT_XLSX_PATH}")


def main():
    report_progress("Loading material codes...")
    mc_norm = load_mc_norm()
    report_progress("Loading dealer info...")
    dealer_info_by_sap, sp_saps, dealer_info_by_code = load_dealer_info()
    report_progress("Loading registration/unavailable VINs and not allocated notes...")
    registration_dates = load_registration_dates()
    registration_vins = set(registration_dates)
    unavailable_vins = load_vin_set_from_table("unavailable", ["vin"], optional=True)
    not_allocated_notes = load_not_allocated_notes()
    not_allocated_ports = load_not_allocated_ports()
    newport_ports = load_newport_ports()
    reservation_vins = load_reservation_vins()
    derogation_rows = load_derogation_rows()
    report_progress("Loading Port Stock ports...")
    port_stock_ports = free_cars_history.load_port_stock_ports(EXCEL_PATHS["in_port"], report_progress)
    report_progress("Building dealer stock pool...")
    stock_rows, vehicle_tracking_derogation = load_dealer_stock_pool(
        mc_norm=mc_norm,
        dealer_info_by_sap=dealer_info_by_sap,
        sp_saps=sp_saps,
        registration_vins=registration_vins,
        unavailable_vins=unavailable_vins,
        derogation_vins={row["vin"] for row in derogation_rows},
        port_stock_ports=port_stock_ports,
        newport_ports=newport_ports,
        not_allocated_ports=not_allocated_ports,
    )
    report_progress("Loading orders...")
    orders = load_orders(mc_norm, dealer_info_by_code)
    report_progress("Applying direct order matches...")
    apply_direct_order_matches(stock_rows, orders)
    report_progress("Applying back-order matches...")
    apply_back_order_matches(stock_rows, orders)
    report_progress("Applying derogations...")
    apply_derogations(stock_rows, registration_vins, derogation_rows)
    report_progress("Applying port stock...")
    apply_port_stock_sources(stock_rows)
    report_progress("Building derogation output...")
    derogation_list_rows = build_derogation_list_rows(
        mc_norm=mc_norm,
        derogation_rows=derogation_rows,
        vehicle_tracking_by_vin=vehicle_tracking_derogation,
        reservation_vins=reservation_vins,
        not_allocated_notes=not_allocated_notes,
        registration_dates=registration_dates,
    )
    report_progress("Writing output...")
    write_output(stock_rows, derogation_list_rows)


if __name__ == "__main__":
    main()

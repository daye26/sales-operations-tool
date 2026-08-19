from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
import re
import sys
import unicodedata
import warnings

from openpyxl import Workbook, load_workbook

import asignaciones_excel as vehicle_tracking_cache
from excel_output import append_row, calculate_column_widths, prepare_worksheet, save_workbook_atomically
import free_cars_history
from port_resolution import resolve_port


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

BASE_EXCEL_DIR = Path.cwd()
OUTPUT_XLSX_PATH = BASE_EXCEL_DIR / "vehicle_preallocation_result.xlsx"
PROGRESS_CALLBACK = None
SHORT_DATE_FORMAT = "yyyy-mm-dd"

EXCEL_PATHS = {
    "reservation": BASE_EXCEL_DIR / "Vehicle_Reservation.xlsx",
    "vehicle_tracking": BASE_EXCEL_DIR / "VehicleTracking.xlsx",
    "mc_norm": BASE_EXCEL_DIR / "material code.xlsx",
    "newport": BASE_EXCEL_DIR / "NEWport.xlsx",
    "not_allocated": BASE_EXCEL_DIR / "Cars not allocated.xlsx",
    "logistics_db": BASE_EXCEL_DIR / "BASE DE DATOS LOGISTICA.xlsx",
}

OUTPUT_COLUMNS = [
    "vin",
    "material_code",
    "description",
    "p_series",
    "int_color",
    "ext_color",
    "status",
    "port",
    "vessel",
    "eta",
    "dealer",
    "or_number",
    "client_order",
    "PREALLOCATION START",
    "days since available",
]

HEADER_ALIASES = {
    "vin": ["vin", "vin #", "vin#"],
    "dealer": ["dealer", "customer name"],
    "or_number": ["or_number", "or number", "order number", "customer reference #"],
    "country": ["country", "customer country"],
    "sales_company": ["sales company"],
    "pre_allocation": [
        "pre_allocation",
        "pre allocation",
        "pre-allocation",
        "preallocation",
        "pre allocation date",
        "preallocation date",
        "reservation date",
    ],
    "code": ["material code", "vehicle material code", "code"],
    "port": ["port", "warehouse", "current warehouse"],
    "p_series": ["product series", "model", "p_series"],
    "int_color": ["interial color", "interior color"],
    "ext_color": ["exterial color", "exterior color"],
}


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def is_missing(value):
    return value is None or str(value).strip() == ""


def report_progress(message):
    if PROGRESS_CALLBACK is None:
        print(message)
        return

    try:
        PROGRESS_CALLBACK(message)
    except Exception:
        pass


def format_value(value):
    if is_missing(value):
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def normalize_header(value):
    text = unicodedata.normalize("NFKC", format_value(value))
    text = text.lower()
    text = text.replace("-", " ").replace("_", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def code_key(value):
    return format_value(value).upper()


def vin_key(value):
    return code_key(value)


def header_index(columns, column_name, required=True):
    aliases = {normalize_header(alias) for alias in HEADER_ALIASES[column_name]}
    for index, column in enumerate(columns):
        if normalize_header(column) in aliases:
            return index
    if required:
        raise ValueError(f"Missing column {column_name}. Headers: {columns}")
    return None


def build_indexes(columns, required_columns, optional_columns=None):
    indexes = {column: header_index(columns, column) for column in required_columns}
    for column in optional_columns or []:
        indexes[column] = header_index(columns, column, required=False)
    return indexes


def row_value(row, indexes, column):
    index = indexes.get(column)
    if index is None or index >= len(row):
        return None
    return row[index]


def open_sheet(key):
    path = EXCEL_PATHS[key]
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    worksheet.reset_dimensions()
    return workbook, worksheet


def read_header(worksheet):
    return [format_value(value) for value in next(worksheet.iter_rows(values_only=True))]


def to_datetime(value):
    if is_missing(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = str(value).strip()
    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass

    return datetime.fromisoformat(text)


def excel_date_value(value):
    parsed = to_datetime(value)
    if parsed is None:
        return ""
    return parsed.date()


def greatest_date(*values):
    parsed = [to_datetime(value) for value in values]
    parsed = [value for value in parsed if value is not None]
    if not parsed:
        return None
    return max(parsed)


def print_warning(title, rows):
    if not rows:
        return

    print(f"WARNING: {title}: {len(rows)}")
    for row in rows:
        print(row)


def load_vehicle_tracking():
    old_path = vehicle_tracking_cache.EXCEL_PATHS["vehicle_tracking"]
    old_callback = vehicle_tracking_cache.PROGRESS_CALLBACK
    try:
        vehicle_tracking_cache.EXCEL_PATHS["vehicle_tracking"] = EXCEL_PATHS["vehicle_tracking"]
        vehicle_tracking_cache.PROGRESS_CALLBACK = PROGRESS_CALLBACK or report_progress
        return vehicle_tracking_cache.load_vehicle_tracking()
    finally:
        vehicle_tracking_cache.EXCEL_PATHS["vehicle_tracking"] = old_path
        vehicle_tracking_cache.PROGRESS_CALLBACK = old_callback


def clear_vehicle_tracking_cache():
    return vehicle_tracking_cache.clear_vehicle_tracking_cache()


def load_port_stock_ports():
    path = EXCEL_PATHS.get("logistics_db")
    if path is None or not Path(path).exists():
        report_progress("WARNING: Logistics database not found; using Vehicle Tracking port")
        return {}
    return free_cars_history.load_port_stock_ports(path, report_progress)


def load_vin_ports(key, label, optional=False):
    path = EXCEL_PATHS.get(key)
    if path is None or not Path(path).exists():
        if optional:
            report_progress(f"WARNING: {label} file not found; using the remaining port sources")
            return {}
        raise FileNotFoundError(path)

    report_progress(f"Loading {label} ports...")
    workbook, worksheet = open_sheet(key)
    try:
        columns = read_header(worksheet)
        try:
            indexes = build_indexes(columns, ["vin", "port"])
        except ValueError:
            if optional:
                report_progress(
                    f"WARNING: {label} does not contain VIN and port columns; "
                    "using the remaining port sources"
                )
                return {}
            raise
        ports_by_vin = {}
        conflicting_vins = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row_value(row, indexes, "vin"))
            port = format_value(row_value(row, indexes, "port"))
            if not vin or not port:
                continue
            previous_port = ports_by_vin.get(vin)
            if previous_port and previous_port != port:
                conflicting_vins.add(vin)
                continue
            ports_by_vin.setdefault(vin, port)
    finally:
        workbook.close()

    if conflicting_vins:
        raise ValueError(f"{label} has conflicting VIN ports: {', '.join(sorted(conflicting_vins)[:20])}")
    report_progress(f"{label} VIN ports loaded: {len(ports_by_vin):,}")
    return ports_by_vin


def load_newport_ports():
    return load_vin_ports("newport", "NEWport", optional=True)


def load_not_allocated_ports():
    return load_vin_ports("not_allocated", "Cars not allocated")


def load_mc_norm():
    report_progress("Loading material codes...")
    workbook, worksheet = open_sheet("mc_norm")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(columns, ["code", "p_series", "int_color", "ext_color"])
        by_code = {}
        counter = Counter()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            code = code_key(row_value(row, indexes, "code"))
            if is_missing(code):
                continue
            counter[code] += 1
            by_code[code] = {
                "p_series": format_value(row_value(row, indexes, "p_series")),
                "int_color": format_value(row_value(row, indexes, "int_color")),
                "ext_color": format_value(row_value(row, indexes, "ext_color")),
            }

        duplicated = [code for code, count in counter.items() if count > 1]
        if duplicated:
            raise ValueError(f"mc_norm duplicated material codes: {', '.join(sorted(duplicated)[:20])}")

        report_progress(f"Material codes loaded: {len(by_code):,}")
        return by_code
    finally:
        workbook.close()


def load_reservations():
    report_progress("Loading Spanish reservations...")
    workbook, worksheet = open_sheet("reservation")
    try:
        columns = read_header(worksheet)
        indexes = build_indexes(
            columns,
            ["vin", "dealer", "or_number", "pre_allocation"],
            optional_columns=["country", "sales_company"],
        )
        if indexes.get("country") is None and indexes.get("sales_company") is None:
            raise ValueError("Reservation file must include either country or Sales Company.")

        rows = []
        warnings_by_type = {
            "missing_vin": [],
        }

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            country = code_key(row_value(row, indexes, "country"))
            sales_company = code_key(row_value(row, indexes, "sales_company"))
            is_spain = country in {"SP", "ES", "SPAIN", "ESPANA", "ESPAÑA"} or sales_company == "9030"
            if not is_spain:
                continue

            vin = vin_key(row_value(row, indexes, "vin"))
            dealer = format_value(row_value(row, indexes, "dealer"))
            or_number = format_value(row_value(row, indexes, "or_number"))
            pre_allocation = row_value(row, indexes, "pre_allocation")
            if not vin:
                warnings_by_type["missing_vin"].append(f"{dealer};{or_number}")
                continue

            rows.append(
                {
                    "vin": vin,
                    "dealer": dealer,
                    "or_number": or_number,
                    "pre_allocation": pre_allocation,
                }
            )

        for title, warning_rows in warnings_by_type.items():
            print_warning(f"reservations {title}", warning_rows)

        report_progress(f"Spanish reservations loaded: {len(rows):,}")
        return rows
    finally:
        workbook.close()


def eta_label(tracking):
    gate_in = to_datetime(tracking.get("gate_in"))
    if gate_in is not None and gate_in.date() <= date.today():
        return "In port"

    eta = excel_date_value(tracking.get("eta"))
    return eta if eta else "In port"


def build_rows(
    reservations,
    vehicle_tracking,
    mc_norm,
    port_stock_ports=None,
    newport_ports=None,
    not_allocated_ports=None,
):
    report_progress("Building vehicle preallocation output...")
    port_stock_ports = port_stock_ports or {}
    newport_ports = newport_ports or {}
    not_allocated_ports = not_allocated_ports or {}
    rows = []
    warnings_by_type = {
        "missing_vehicle_tracking": [],
        "missing_mc_norm": [],
    }

    for reservation in reservations:
        vin = reservation["vin"]
        tracking = vehicle_tracking.get(vin)
        if tracking is None:
            warnings_by_type["missing_vehicle_tracking"].append(
                f"{vin};{reservation['dealer']};{reservation['or_number']}"
            )
            tracking = {}

        material_code = code_key(tracking.get("material_code"))
        mc = mc_norm.get(material_code, {})
        if material_code and not mc:
            warnings_by_type["missing_mc_norm"].append(f"{vin};{material_code}")

        gate_in = to_datetime(tracking.get("gate_in"))
        preallocation_start = (
            greatest_date(reservation["pre_allocation"], gate_in)
            if gate_in is not None
            else None
        )
        or_number = reservation["or_number"]
        rows.append(
            (
                vin,
                material_code,
                format_value(tracking.get("description")),
                mc.get("p_series", ""),
                mc.get("int_color", ""),
                mc.get("ext_color", ""),
                format_value(tracking.get("status")),
                resolve_port(
                    port_stock_ports.get(vin),
                    newport_ports.get(vin),
                    not_allocated_ports.get(vin),
                    tracking.get("port"),
                ),
                format_value(tracking.get("vessel_name")),
                eta_label(tracking),
                reservation["dealer"],
                or_number,
                "Y" if or_number.upper().startswith("OR") else "",
                excel_date_value(preallocation_start),
                (date.today() - preallocation_start.date()).days
                if preallocation_start is not None
                else "",
            )
        )

    for title, warning_rows in warnings_by_type.items():
        print_warning(title, warning_rows)

    report_progress(f"Vehicle preallocation rows built: {len(rows):,}")
    return rows


def write_output(rows):
    report_progress("Writing output workbook...")
    workbook = Workbook(write_only=True)
    try:
        worksheet = workbook.create_sheet("VEHICLE_PREALLOCATION")
        row_values = lambda row: list(row)
        column_widths = calculate_column_widths(OUTPUT_COLUMNS, rows, row_values, format_value)
        prepare_worksheet(worksheet, OUTPUT_COLUMNS, column_widths, 45)
        for row in rows:
            append_row(worksheet, row_values(row), SHORT_DATE_FORMAT)
        save_workbook_atomically(workbook, OUTPUT_XLSX_PATH)
    finally:
        workbook.close()

    print("PROCESS_SUMMARY")
    print(f"vehicle_preallocation_rows: {len(rows)}")
    print(f"client_order_y: {sum(1 for row in rows if row[12] == 'Y')}")
    print(f"client_order_blank: {sum(1 for row in rows if row[12] != 'Y')}")
    print(f"OUTPUT_XLSX: {OUTPUT_XLSX_PATH}")


def main():
    report_progress("Starting vehicle preallocation...")
    mc_norm = load_mc_norm()
    vehicle_tracking = load_vehicle_tracking()
    port_stock_ports = load_port_stock_ports()
    newport_ports = load_newport_ports()
    not_allocated_ports = load_not_allocated_ports()
    reservations = load_reservations()
    rows = build_rows(
        reservations,
        vehicle_tracking,
        mc_norm,
        port_stock_ports,
        newport_ports,
        not_allocated_ports,
    )
    write_output(rows)


if __name__ == "__main__":
    main()

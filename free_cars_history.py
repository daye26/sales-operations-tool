"""Persistent recovery history for Cars not allocated and Port Stock records."""

from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
import os
import re
import sqlite3
import unicodedata

from openpyxl import load_workbook


DATABASE_FILENAME = "sales_operations_history.sqlite"
PORT_STOCK_SHEET_NAME = "Port Stock"
PORT_STOCK_SOURCE_KEY = "logistics_port_stock"


def history_database_path():
    appdata = os.environ.get("APPDATA")
    base = Path(appdata) if appdata else Path.home() / ".sales_operations_tool"
    return base / "Sales Operations Tool" / "data" / DATABASE_FILENAME


def vin_key(value):
    return "" if value is None else str(value).strip().upper()


def value_text(value):
    return "" if value is None else str(value).strip()


def normalize_header(value):
    text = unicodedata.normalize("NFKC", value_text(value)).lower()
    return " ".join(text.replace("-", " ").replace("_", " ").split())


def match_group_letters(number):
    letters = []
    while number:
        number, remainder = divmod(number - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def normalize_match_group_value(value):
    """Render valid legacy numeric groups as their letter equivalents."""
    text = value_text(value)
    if not text:
        return ""

    groups = []
    seen = set()
    for raw_group in text.split(";"):
        group = raw_group.strip().upper()
        if not re.fullmatch(r"[A-Z]+", group):
            try:
                group_number = Decimal(raw_group.strip())
            except InvalidOperation:
                return text
            if (
                not group_number.is_finite()
                or group_number < 1
                or group_number != group_number.to_integral_value()
            ):
                return text
            group = match_group_letters(int(group_number))
        if group not in seen:
            groups.append(group)
            seen.add(group)
    return ";".join(groups)


def table_exists(connection, table_name):
    return connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone() is not None


def ensure_schema(connection):
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS free_cars_history (
            vin TEXT PRIMARY KEY NOT NULL,
            priority TEXT,
            match_group TEXT,
            note TEXT,
            warehouse TEXT,
            delete_date TEXT
        )
        """
    )
    columns = {row[1] for row in connection.execute("PRAGMA table_info(free_cars_history)")}
    for column_name in ("priority", "match_group", "note", "warehouse", "delete_date"):
        if column_name not in columns:
            connection.execute(f"ALTER TABLE free_cars_history ADD COLUMN {column_name} TEXT")
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS port_stock_history (
            vin TEXT PRIMARY KEY NOT NULL,
            port TEXT NOT NULL
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS history_source_state (
            source_key TEXT PRIMARY KEY NOT NULL,
            source_path TEXT NOT NULL,
            source_size INTEGER NOT NULL,
            source_mtime_ns INTEGER NOT NULL
        )
        """
    )


def sync_not_allocated_records(records, database_path=None, today=None):
    """Keep the final known Cars not allocated values for every VIN.

    Active registration VINs are deliberately excluded: registration is the
    authoritative signal that a VIN no longer belongs in this recovery history.
    """
    database_path = Path(database_path) if database_path else history_database_path()
    database_path.parent.mkdir(parents=True, exist_ok=True)
    delete_date = (today or date.today()).isoformat()

    connection = sqlite3.connect(database_path)
    try:
        ensure_schema(connection)
        connection.execute(
            """
            CREATE TEMP TABLE current_free_cars (
                vin TEXT PRIMARY KEY NOT NULL,
                priority TEXT,
                match_group TEXT,
                note TEXT,
                warehouse TEXT
            )
            """
        )

        rows_buffer = []
        scanned = 0
        for record in records:
            vin = vin_key(record.get("vin"))
            if not vin:
                continue
            scanned += 1
            rows_buffer.append(
                (
                    vin,
                    value_text(record.get("priority")),
                    normalize_match_group_value(record.get("match_group")),
                    value_text(record.get("note")),
                    value_text(record.get("warehouse")),
                )
            )
            if len(rows_buffer) == 1000:
                _upsert_current_rows(connection, rows_buffer)
                rows_buffer.clear()
        if rows_buffer:
            _upsert_current_rows(connection, rows_buffer)

        registered_removed = remove_registered_vins(connection=connection)
        if table_exists(connection, "registration_history"):
            connection.execute(
                """
                DELETE FROM current_free_cars
                WHERE EXISTS (
                    SELECT 1
                    FROM registration_history r
                    WHERE r.vin = current_free_cars.vin
                      AND r.is_active = 1
                )
                """
            )

        new_rows = connection.execute(
            """
            SELECT COUNT(*)
            FROM current_free_cars c
            WHERE NOT EXISTS (
                SELECT 1 FROM free_cars_history h WHERE h.vin = c.vin
            )
            """
        ).fetchone()[0]
        reactivated = connection.execute(
            """
            SELECT COUNT(*)
            FROM current_free_cars c
            INNER JOIN free_cars_history h ON h.vin = c.vin
            WHERE h.delete_date IS NOT NULL
            """
        ).fetchone()[0]
        marked_deleted = connection.execute(
            """
            SELECT COUNT(*)
            FROM free_cars_history h
            WHERE h.delete_date IS NULL
              AND NOT EXISTS (
                  SELECT 1 FROM current_free_cars c WHERE c.vin = h.vin
              )
            """
        ).fetchone()[0]

        connection.execute(
            """
            UPDATE free_cars_history
            SET delete_date = ?
            WHERE delete_date IS NULL
              AND NOT EXISTS (
                  SELECT 1 FROM current_free_cars c WHERE c.vin = free_cars_history.vin
              )
            """,
            (delete_date,),
        )
        connection.execute(
            """
            UPDATE free_cars_history
            SET priority = (
                    SELECT c.priority FROM current_free_cars c WHERE c.vin = free_cars_history.vin
                ),
                match_group = (
                    SELECT c.match_group FROM current_free_cars c WHERE c.vin = free_cars_history.vin
                ),
                note = (
                    SELECT c.note FROM current_free_cars c WHERE c.vin = free_cars_history.vin
                ),
                warehouse = (
                    SELECT c.warehouse FROM current_free_cars c WHERE c.vin = free_cars_history.vin
                ),
                delete_date = NULL
            WHERE EXISTS (
                SELECT 1 FROM current_free_cars c WHERE c.vin = free_cars_history.vin
            )
            """
        )
        connection.execute(
            """
            INSERT INTO free_cars_history (vin, priority, match_group, note, warehouse, delete_date)
            SELECT c.vin, c.priority, c.match_group, c.note, c.warehouse, NULL
            FROM current_free_cars c
            WHERE NOT EXISTS (
                SELECT 1 FROM free_cars_history h WHERE h.vin = c.vin
            )
            """
        )
        active_rows = connection.execute(
            "SELECT COUNT(*) FROM current_free_cars"
        ).fetchone()[0]
        connection.commit()
    finally:
        connection.close()

    return {
        "database_path": database_path,
        "scanned": scanned,
        "active_rows": active_rows,
        "new": new_rows,
        "reactivated": reactivated,
        "marked_deleted": marked_deleted,
        "registered_removed": registered_removed,
    }


def _upsert_current_rows(connection, rows):
    connection.executemany(
        """
        INSERT INTO current_free_cars (vin, priority, match_group, note, warehouse)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(vin) DO UPDATE SET
            priority = excluded.priority,
            match_group = excluded.match_group,
            note = excluded.note,
            warehouse = excluded.warehouse
        """,
        rows,
    )


def remove_registered_vins(database_path=None, connection=None):
    """Hard-delete Free Cars history rows already registered in ANT."""
    owns_connection = connection is None
    if owns_connection:
        database_path = Path(database_path) if database_path else history_database_path()
        if not database_path.exists():
            return 0
        connection = sqlite3.connect(database_path)

    try:
        ensure_schema(connection)
        if not table_exists(connection, "registration_history"):
            return 0
        cursor = connection.execute(
            """
            DELETE FROM free_cars_history
            WHERE EXISTS (
                SELECT 1
                FROM registration_history r
                WHERE r.vin = free_cars_history.vin
                  AND r.is_active = 1
            )
            """
        )
        if owns_connection:
            connection.commit()
        return cursor.rowcount
    finally:
        if owns_connection:
            connection.close()


def load_deleted_records(database_path=None):
    """Return retained values only for VINs no longer in Cars not allocated."""
    database_path = Path(database_path) if database_path else history_database_path()
    if not database_path.exists():
        return {}

    connection = sqlite3.connect(database_path)
    try:
        ensure_schema(connection)
        rows = connection.execute(
            """
            SELECT vin, priority, match_group, note, warehouse, delete_date
            FROM free_cars_history
            WHERE delete_date IS NOT NULL
            """
        ).fetchall()
    finally:
        connection.close()

    return {
        vin: {
            "priority": value_text(priority),
            "match_group": normalize_match_group_value(match_group),
            "note": value_text(note),
            "warehouse": value_text(warehouse),
            "delete_date": value_text(delete_date),
        }
        for vin, priority, match_group, note, warehouse, delete_date in rows
    }


def port_stock_signature(source_path):
    source_path = Path(source_path)
    stat = source_path.stat()
    return {
        "path": os.path.normcase(str(source_path.resolve())),
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
    }


def _report(report_progress, message):
    if report_progress is not None:
        report_progress(message)


def _read_port_stock_records(source_path, report_progress):
    _report(report_progress, "Reading Logistics database / Port Stock ports...")
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if PORT_STOCK_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Logistics database is missing sheet {PORT_STOCK_SHEET_NAME!r}. "
                f"Available sheets: {workbook.sheetnames}"
            )
        worksheet = workbook[PORT_STOCK_SHEET_NAME]
        worksheet.reset_dimensions()
        try:
            headers = list(next(worksheet.iter_rows(max_row=1, values_only=True)))
        except StopIteration as exc:
            raise ValueError(f"Logistics database sheet {PORT_STOCK_SHEET_NAME!r} is empty.") from exc

        header_indexes = {normalize_header(header): index for index, header in enumerate(headers)}
        vin_index = header_indexes.get("vin")
        port_index = header_indexes.get("port")
        if vin_index is None or port_index is None:
            raise ValueError(
                "Logistics database / Port Stock requires VIN and PORT columns. "
                f"Headers: {headers}"
            )

        ports_by_vin = {}
        duplicated_vins = set()
        skipped_rows = 0
        max_col = max(vin_index, port_index) + 1
        for row in worksheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
            if not any(value_text(value) for value in row):
                continue
            vin = vin_key(row[vin_index] if vin_index < len(row) else None)
            port = value_text(row[port_index] if port_index < len(row) else None)
            if not vin or not port:
                skipped_rows += 1
                continue
            if vin in ports_by_vin:
                duplicated_vins.add(vin)
                continue
            ports_by_vin[vin] = port
    finally:
        workbook.close()

    if duplicated_vins:
        examples = "; ".join(sorted(duplicated_vins)[:20])
        raise ValueError(f"Port Stock has duplicated VINs: {examples}")
    if skipped_rows:
        _report(report_progress, f"WARNING: Port Stock skipped {skipped_rows:,} row(s) without VIN or PORT")
    if not ports_by_vin:
        raise ValueError("Port Stock has no VINs with PORT values.")
    return ports_by_vin


def load_port_stock_ports(source_path, report_progress=None, database_path=None):
    """Return historical Port Stock ports, refreshing only when its Excel changes.

    Removed VINs deliberately remain available with their last known PORT. A later
    reappearance updates that PORT again.
    """
    source_path = Path(source_path)
    if not source_path.exists():
        raise FileNotFoundError(source_path)

    database_path = Path(database_path) if database_path else history_database_path()
    database_path.parent.mkdir(parents=True, exist_ok=True)
    signature = port_stock_signature(source_path)
    connection = sqlite3.connect(database_path)
    try:
        ensure_schema(connection)
        state = connection.execute(
            """
            SELECT source_path, source_size, source_mtime_ns
            FROM history_source_state
            WHERE source_key = ?
            """,
            (PORT_STOCK_SOURCE_KEY,),
        ).fetchone()
        source_unchanged = state == (
            signature["path"],
            signature["size"],
            signature["mtime_ns"],
        )
        if not source_unchanged:
            current_ports = _read_port_stock_records(source_path, report_progress)
            connection.executemany(
                """
                INSERT INTO port_stock_history (vin, port)
                VALUES (?, ?)
                ON CONFLICT(vin) DO UPDATE SET port = excluded.port
                """,
                current_ports.items(),
            )
            connection.execute(
                """
                INSERT INTO history_source_state (
                    source_key, source_path, source_size, source_mtime_ns
                ) VALUES (?, ?, ?, ?)
                ON CONFLICT(source_key) DO UPDATE SET
                    source_path = excluded.source_path,
                    source_size = excluded.source_size,
                    source_mtime_ns = excluded.source_mtime_ns
                """,
                (
                    PORT_STOCK_SOURCE_KEY,
                    signature["path"],
                    signature["size"],
                    signature["mtime_ns"],
                ),
            )
            connection.commit()
            _report(
                report_progress,
                f"Port Stock history synchronized: {len(current_ports):,} current VINs retained",
            )

        rows = connection.execute("SELECT vin, port FROM port_stock_history").fetchall()
    finally:
        connection.close()

    ports_by_vin = {vin_key(vin): value_text(port) for vin, port in rows}
    source_status = "hit" if source_unchanged else "refreshed"
    _report(report_progress, f"Port Stock history {source_status}: {len(ports_by_vin):,} VIN ports available")
    return ports_by_vin

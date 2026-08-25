from calendar import month_name
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path
import re
import sys
import unicodedata
import warnings

from openpyxl import Workbook, load_workbook

import asignaciones_excel as preallocation_engine
from excel_sheet_selection import select_active_then_sheet1
import free_cars_history
from excel_output import append_row, calculate_column_widths, prepare_worksheet, save_workbook_atomically
from port_resolution import resolve_port
import tabular_normalization as tabular


warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

BASE_EXCEL_DIR = Path.cwd()
OUTPUT_XLSX_PATH = BASE_EXCEL_DIR / "check_free_cars_result.xlsx"
PROGRESS_CALLBACK = None
SHORT_DATE_FORMAT = "yyyy-mm-dd"

EXCEL_PATHS = {
    "vehicle_tracking": BASE_EXCEL_DIR / "VehicleTracking.xlsx",
    "mc_norm": BASE_EXCEL_DIR / "material code.xlsx",
    "not_allocated": BASE_EXCEL_DIR / "Cars not allocated.xlsx",
    "newport": BASE_EXCEL_DIR / "NEWport.xlsx",
    "reservation": BASE_EXCEL_DIR / "Vehicle_Reservation.xlsx",
    "unavailable": BASE_EXCEL_DIR / "unavailable.xlsx",
    "logistics_db": BASE_EXCEL_DIR / "BASE DE DATOS LOGISTICA.xlsx",
}
SHEET_NAMES = {
    "not_allocated": "Sheet1",
}

HEADER_ALIASES = {
    "vin": ["vin", "vin #"],
    "material_code": ["material code", "vehicle material code", "code"],
    "note": ["note"],
    "month": ["month"],
    "port": ["port", "current warehouse"],
    "priority": ["priority", "\u4f18\u5148\u5206\u914d"],
    "match_group": ["match group", "match_group"],
    "warehouse": [
        "\u9a6c\u5fb7\u91cc&\u5df4\u585e&\u74e6\u4f26\u4ed3\u5e93",
        "madrid barcelona valencia warehouse",
        "\u9a6c\u5fb7\u91cc&\u5df4\u585e\u4ed3\u5e93",
        "madrid barcelona warehouse",
        "warehouse",
    ],
    "code": ["material code", "vehicle material code", "code"],
    "p_series": ["product series", "model", "p_series"],
    "p_trim": ["product model", "trim", "p_trim"],
    "int_color": ["interial color", "interior color", "int_color"],
    "ext_color": ["exterial color", "exterior color", "ext_color"],
    "model_year": ["model year", "model_year", "my"],
}


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def report_progress(message):
    if PROGRESS_CALLBACK is None:
        print(message)
        return

    try:
        PROGRESS_CALLBACK(message)
    except Exception:
        pass


is_missing = tabular.is_missing
format_value = tabular.format_value
normalize_header = tabular.normalize_header
text_key = tabular.text_key
vin_key = tabular.vin_key


def to_datetime(value):
    if is_missing(value):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = format_value(value)
    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass
    return datetime.fromisoformat(text)


def to_date(value):
    parsed = to_datetime(value)
    return parsed.date() if parsed is not None else ""


def header_index(headers, field, required=True):
    return tabular.header_index(headers, HEADER_ALIASES, field, required)


row_value = tabular.row_value


def open_sheet(key):
    path = EXCEL_PATHS[key]
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    selected_sheet = SHEET_NAMES.get(key)
    if key == "not_allocated":
        worksheet = select_active_then_sheet1(
            workbook,
            HEADER_ALIASES,
            ("vin", "material_code", "note"),
        )
    else:
        worksheet = workbook[selected_sheet] if selected_sheet in workbook.sheetnames else workbook.active
    worksheet.reset_dimensions()
    return workbook, worksheet


def read_header(worksheet):
    try:
        return list(next(worksheet.iter_rows(values_only=True)))
    except StopIteration as exc:
        raise ValueError(f"The worksheet {worksheet.title} is empty.") from exc


def load_not_allocated():
    report_progress("Loading Cars not allocated...")
    workbook, worksheet = open_sheet("not_allocated")
    try:
        raw_headers = read_header(worksheet)
        indexes = {
            field: header_index(raw_headers, field)
            for field in ("vin", "note", "month", "port", "material_code")
        }
        for field in ("priority", "match_group", "warehouse"):
            indexes[field] = header_index(raw_headers, field, required=False)
        records = []
        vin_counter = Counter()

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row_value(row, indexes, "vin"))
            if not vin:
                continue
            vin_counter[vin] += 1
            records.append(
                {
                    "vin": vin,
                    "note": row_value(row, indexes, "note"),
                    "month": row_value(row, indexes, "month"),
                    "port": row_value(row, indexes, "port"),
                    "material_code": format_value(row_value(row, indexes, "material_code")).upper(),
                    "priority": format_value(row_value(row, indexes, "priority")),
                    "match_group": preallocation_engine.format_match_groups(
                        preallocation_engine.parse_match_groups(
                            row_value(row, indexes, "match_group"),
                            f"Cars not allocated Match group row {row_number}",
                        )
                    ),
                    "warehouse": format_value(row_value(row, indexes, "warehouse")),
                }
            )

        duplicated = sum(count - 1 for count in vin_counter.values() if count > 1)
        report_progress(
            f"Cars not allocated loaded: {len(records):,} rows, {duplicated:,} additional duplicated VIN row(s)"
        )
        return records
    finally:
        workbook.close()


def load_mc_norm():
    report_progress("Loading material codes...")
    workbook, worksheet = open_sheet("mc_norm")
    try:
        raw_headers = read_header(worksheet)
        indexes = {
            field: header_index(raw_headers, field)
            for field in ("code", "p_series", "p_trim", "int_color", "ext_color", "model_year")
        }
        by_code = {}
        counter = Counter()

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            code = format_value(row_value(row, indexes, "code")).upper()
            if not code:
                continue
            counter[code] += 1
            p_series = format_value(row_value(row, indexes, "p_series"))
            p_trim = format_value(row_value(row, indexes, "p_trim"))
            int_color = format_value(row_value(row, indexes, "int_color"))
            ext_color = format_value(row_value(row, indexes, "ext_color"))
            by_code[code] = {
                "p_series": p_series,
                "p_trim": p_trim,
                "int_color": int_color,
                "ext_color": ext_color,
                "model_year": format_value(row_value(row, indexes, "model_year")),
                "aggr": f"{p_series}{p_trim}{int_color}{ext_color}",
                "raw": tuple(row),
            }

        duplicated = sorted(code for code, count in counter.items() if count > 1)
        if duplicated:
            raise ValueError(f"mc_norm duplicated material codes: {', '.join(duplicated[:10])}")

        report_progress(f"Material codes loaded: {len(by_code):,}")
        return by_code
    finally:
        workbook.close()


def load_vin_set(key, label):
    report_progress(f"Loading {label} VINs...")
    workbook, worksheet = open_sheet(key)
    try:
        headers = read_header(worksheet)
        vin_index = header_index(headers, "vin")
        vins = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row[vin_index] if vin_index < len(row) else None)
            if vin:
                vins.add(vin)

        report_progress(f"{label} VINs loaded: {len(vins):,}")
        return vins
    finally:
        workbook.close()


def load_vehicle_tracking():
    old_path = preallocation_engine.EXCEL_PATHS["vehicle_tracking"]
    old_callback = preallocation_engine.PROGRESS_CALLBACK
    try:
        preallocation_engine.EXCEL_PATHS["vehicle_tracking"] = EXCEL_PATHS["vehicle_tracking"]
        preallocation_engine.PROGRESS_CALLBACK = PROGRESS_CALLBACK or report_progress
        return preallocation_engine.load_vehicle_tracking()
    finally:
        preallocation_engine.EXCEL_PATHS["vehicle_tracking"] = old_path
        preallocation_engine.PROGRESS_CALLBACK = old_callback


def load_port_stock_ports():
    path = EXCEL_PATHS.get("logistics_db")
    if path is None or not Path(path).exists():
        report_progress("WARNING: Logistics database not found; using Vehicle Tracking port")
        return {}
    return free_cars_history.load_port_stock_ports(path, report_progress)


def load_newport_ports():
    path = EXCEL_PATHS.get("newport")
    if path is None or not Path(path).exists():
        report_progress("WARNING: NEWport file not found; using the remaining port sources")
        return {}

    report_progress("Loading NEWport ports...")
    workbook, worksheet = open_sheet("newport")
    try:
        headers = read_header(worksheet)
        try:
            indexes = {
                "vin": header_index(headers, "vin"),
                "port": header_index(headers, "port"),
            }
        except ValueError:
            report_progress(
                "WARNING: NEWport does not contain VIN and port columns; "
                "using the remaining port sources"
            )
            return {}
        ports_by_vin = {}
        duplicate_vins = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(not is_missing(value) for value in row):
                continue
            vin = vin_key(row_value(row, indexes, "vin"))
            port = format_value(row_value(row, indexes, "port"))
            if not vin or not port:
                raise ValueError("NEWport has a row with an empty VIN or port")
            if vin in ports_by_vin:
                duplicate_vins.add(vin)
            ports_by_vin[vin] = port
        if duplicate_vins:
            raise ValueError("NEWport has duplicated VINs: " + ", ".join(sorted(duplicate_vins)[:20]))
        report_progress(f"NEWport VIN ports loaded: {len(ports_by_vin):,}")
        return ports_by_vin
    finally:
        workbook.close()


def is_offline(value):
    return text_key(value) == "OFFLINE"


def is_non_offline_sql(value):
    return not is_missing(value) and not is_offline(value)


def unique_rows(rows):
    seen = set()
    unique = []
    for row in rows:
        if row in seen:
            continue
        seen.add(row)
        unique.append(row)
    return unique


def build_core_outputs(
    not_allocated_records,
    vehicle_tracking,
    mc_norm,
    reservation_vins,
    unavailable_vins,
    deleted_history_by_vin=None,
    port_stock_ports=None,
    newport_ports=None,
):
    not_allocated_vins = {record["vin"] for record in not_allocated_records}
    deleted_history_by_vin = deleted_history_by_vin or {}
    port_stock_ports = port_stock_ports or {}
    newport_ports = newport_ports or {}

    assigned_rows = []
    for vin in sorted(vehicle_tracking):
        tracking = vehicle_tracking[vin]
        if (
            not is_missing(tracking.get("dsn"))
            or not is_missing(tracking.get("reserved_so"))
            or vin in not_allocated_vins
            or vin in unavailable_vins
            or vin in reservation_vins
            or not is_non_offline_sql(tracking.get("status"))
        ):
            continue

        material = mc_norm.get(format_value(tracking.get("material_code")).upper(), {})
        history = deleted_history_by_vin.get(vin, {})
        eta = to_date(tracking.get("eta"))
        gate_in = to_date(tracking.get("gate_in"))
        assigned_rows.append(
            (
                format_value(history.get("note")) or format_value(tracking.get("tag")),
                month_name[eta.month] if is_missing(gate_in) and eta else "",
                "",
                format_value(history.get("priority")),
                resolve_port(
                    port_stock_ports.get(vin),
                    newport_ports.get(vin),
                    None,
                    tracking.get("port"),
                ),
                format_value(history.get("warehouse")),
                format_value(history.get("match_group")),
                vin,
                format_value(tracking.get("material_code")),
                format_value(tracking.get("description")),
                material.get("int_color", ""),
                material.get("ext_color", ""),
                material.get("p_series", ""),
                material.get("p_trim", ""),
                eta,
                format_value(tracking.get("vessel_name")),
                material.get("model_year", ""),
                to_date(tracking.get("production_date")),
            )
        )

    note_tag_rows = []
    reserved_or_allocated_rows = []
    got_on_boat_rows = []
    not_offline_rows = []
    port_mismatch_rows = []
    seen_reserved_or_allocated_vins = set()

    for record in not_allocated_records:
        vin = record["vin"]
        port_stock_port = format_value(port_stock_ports.get(vin))
        not_allocated_port = format_value(record.get("port"))
        if port_stock_port and text_key(not_allocated_port) != text_key(port_stock_port):
            port_mismatch_rows.append(
                (
                    vin,
                    not_allocated_port,
                    port_stock_port,
                )
            )

        tracking = vehicle_tracking.get(vin)
        if tracking is None:
            continue
        material = mc_norm.get(format_value(tracking.get("material_code")).upper(), {})
        note_is_null = is_missing(record["note"])
        tag_is_null = is_missing(tracking.get("tag"))

        if note_is_null != tag_is_null:
            note_tag_rows.append(
                (
                    vin,
                    format_value(record["note"]),
                    format_value(tracking.get("tag")),
                    format_value(tracking.get("material_code")),
                    material.get("aggr", ""),
                )
            )

        has_dsn_data = any(
            not is_missing(tracking.get(field))
            for field in ("dsn", "reserved_so", "related_order")
        )
        if has_dsn_data and vin not in seen_reserved_or_allocated_vins:
            seen_reserved_or_allocated_vins.add(vin)
            reserved_or_allocated_rows.append(
                (
                    vin,
                    format_value(record["note"]),
                    format_value(tracking.get("tag")),
                    format_value(tracking.get("dsn")),
                    format_value(tracking.get("reserved_so")),
                    to_date(tracking.get("dn_create_time")),
                    "Y" if vin in reservation_vins else "",
                )
            )

        month = format_value(record["month"])
        if (
            month
            and
            (is_offline(tracking.get("status")) or is_missing(tracking.get("vessel_name")))
            and "未发船" not in month
            and "未发运" not in month
        ):
            got_on_boat_rows.append(
                (
                    vin,
                    format_value(record["note"]),
                    month,
                    resolve_port(
                        port_stock_ports.get(vin),
                        newport_ports.get(vin),
                        record.get("port"),
                        tracking.get("port"),
                    ),
                    format_value(tracking.get("vessel_name")),
                    to_date(tracking.get("eta")),
                )
            )

        if is_non_offline_sql(tracking.get("status")) and (
            "未发船" in month or "未发运" in month
        ):
            not_offline_rows.append(
                (
                    vin,
                    to_date(tracking.get("eta")),
                    resolve_port(
                        port_stock_ports.get(vin),
                        newport_ports.get(vin),
                        record.get("port"),
                        tracking.get("port"),
                    ),
                    month,
                )
            )

    for record in not_allocated_records:
        vin = record["vin"]
        if vin not in reservation_vins or vin in seen_reserved_or_allocated_vins:
            continue
        seen_reserved_or_allocated_vins.add(vin)
        tracking = vehicle_tracking.get(vin, {})
        reserved_or_allocated_rows.append(
            (
                vin,
                format_value(record["note"]),
                format_value(tracking.get("tag")),
                format_value(tracking.get("dsn")),
                format_value(tracking.get("reserved_so")),
                to_date(tracking.get("dn_create_time")),
                "Y",
            )
        )

    return {
        "assigned": unique_rows(assigned_rows),
        "note_tag": unique_rows(note_tag_rows),
        "reserved_or_allocated": unique_rows(reserved_or_allocated_rows),
        "got_on_boat": unique_rows(got_on_boat_rows),
        "not_offline": unique_rows(not_offline_rows),
        "port_mismatches": unique_rows(port_mismatch_rows),
    }


def clean_output_value(value):
    if isinstance(value, datetime):
        return value.date()
    return value


def write_sheet(workbook, sheet_name, headers, rows):
    if workbook.write_only:
        worksheet = workbook.create_sheet()
    else:
        worksheet = (
            workbook.active
            if len(workbook.worksheets) == 1 and workbook.active.max_row == 1
            else workbook.create_sheet()
        )
    worksheet.title = sheet_name
    row_values = lambda row: [clean_output_value(value) for value in row]
    column_widths = calculate_column_widths(headers, rows, row_values, format_value)
    prepare_worksheet(worksheet, headers, column_widths, 42)
    for row in rows:
        append_row(worksheet, row_values(row), SHORT_DATE_FORMAT)


def write_output(outputs):
    report_progress("Writing Check Free Cars workbook...")
    workbook = Workbook(write_only=True)
    try:
        write_sheet(
            workbook,
            "Available cars",
            [
                "note",
                "month",
                "blank",
                "priority",
                "port",
                "warehouse",
                "match_group",
                "vin",
                "material_code",
                "description",
                "int_color",
                "ext_color",
                "p_series",
                "p_trim",
                "eta",
                "vessel",
                "model_year",
                "production_date",
            ],
            outputs["assigned"],
        )
        write_sheet(
            workbook,
            "Notes = tag",
            ["vin", "note", "tag", "material_code", "aggr"],
            outputs["note_tag"],
        )
        write_sheet(
            workbook,
            "Reserved or allocated cars",
            ["vin", "note", "tag", "dsn", "reserved_so", "dn_create_time", "preallocated"],
            outputs["reserved_or_allocated"],
        )
        write_sheet(
            workbook,
            "Not shipped",
            ["vin", "note", "month", "port", "vessel", "eta"],
            outputs["got_on_boat"],
        )
        write_sheet(
            workbook,
            "No longer offline",
            ["vin", "eta", "port", "month"],
            outputs["not_offline"],
        )
        write_sheet(
            workbook,
            "Port mismatches",
            ["vin", "not_allocated_port", "port_stock_port"],
            outputs.get("port_mismatches", []),
        )
        save_workbook_atomically(workbook, OUTPUT_XLSX_PATH)
    finally:
        workbook.close()

    print("PROCESS_SUMMARY")
    print(f"assigned_cars: {len(outputs['assigned'])}")
    print(f"notes_tag_mismatches: {len(outputs['note_tag'])}")
    print(f"reserved_or_allocated_cars: {len(outputs['reserved_or_allocated'])}")
    print(f"not_shipped: {len(outputs['got_on_boat'])}")
    print(f"no_longer_offline: {len(outputs['not_offline'])}")
    print(f"port_mismatches: {len(outputs.get('port_mismatches', []))}")
    print(f"OUTPUT_XLSX: {OUTPUT_XLSX_PATH}")


def main():
    report_progress("Starting Check Free Cars...")
    not_allocated_records = load_not_allocated()
    mc_norm = load_mc_norm()
    reservation_vins = load_vin_set("reservation", "Reservation")
    unavailable_vins = load_vin_set("unavailable", "Unavailable")
    vehicle_tracking = load_vehicle_tracking()
    port_stock_ports = load_port_stock_ports()
    newport_ports = load_newport_ports()
    history_sync = free_cars_history.sync_not_allocated_records(not_allocated_records)
    report_progress(
        "Free Cars history synchronized: "
        f"new {history_sync['new']:,}, reactivated {history_sync['reactivated']:,}, "
        f"marked deleted {history_sync['marked_deleted']:,}, "
        f"registered removed {history_sync['registered_removed']:,}"
    )
    deleted_history_by_vin = free_cars_history.load_deleted_records()
    outputs = build_core_outputs(
        not_allocated_records,
        vehicle_tracking,
        mc_norm,
        reservation_vins,
        unavailable_vins,
        deleted_history_by_vin,
        port_stock_ports,
        newport_ports,
    )
    write_output(outputs)
    report_progress("Check Free Cars finished")


if __name__ == "__main__":
    main()

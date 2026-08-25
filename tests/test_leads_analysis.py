import csv
from collections import Counter
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook

import leads_analysis_excel as engine
import sales_operations_app as app


def write_csv(path, headers, rows=()):
    with path.open("w", encoding="utf-8-sig", newline="") as source_file:
        writer = csv.writer(source_file)
        writer.writerow(headers)
        writer.writerows(rows)


def write_model_equivalences(path, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["id", "model"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


class LeadsAnalysisTests(unittest.TestCase):
    def setUp(self):
        self.old_paths = engine.EXCEL_PATHS.copy()
        self.old_output = engine.OUTPUT_XLSX_PATH

    def tearDown(self):
        engine.EXCEL_PATHS.clear()
        engine.EXCEL_PATHS.update(self.old_paths)
        engine.OUTPUT_XLSX_PATH = self.old_output

    def test_formula_accepts_only_supported_arithmetic(self):
        _, expression = engine.parse_test_drive_formula(
            "offer_request + dealer_leads * 0.5 + test_drive_request / 2"
        )
        value = engine.calculate_metric_value(
            Counter(offer_request=1, dealer_leads=4, test_drive_request=3),
            "test_drive",
            expression,
        )
        self.assertEqual(5, value)
        with self.assertRaisesRegex(ValueError, "Unknown test drive variable"):
            engine.parse_test_drive_formula("unknown_metric + 1")
        with self.assertRaisesRegex(ValueError, "only supports"):
            engine.parse_test_drive_formula("round(dealer_leads)")

    def test_report_includes_configured_models_without_activity(self):
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            headers = [
                "Create Date",
                "Request Type",
                "Interest Model",
                "Interest Model (External)",
            ]
            leads_sp_path = temp_path / "leads_sp.csv"
            leads_pt_path = temp_path / "leads_pt.csv"
            model_equivalence_path = temp_path / "model_equivalence.xlsx"
            output_path = temp_path / "leads_output.xlsx"
            write_csv(leads_sp_path, headers)
            write_csv(leads_pt_path, headers)
            write_model_equivalences(
                model_equivalence_path,
                [
                    ("ID_Z", "Z Model"),
                    ("ID_A", "A Model"),
                    ("ID_A_DUPLICATE", " a model "),
                    ("ID_OLD", "OLD"),
                ],
            )

            engine.EXCEL_PATHS.update(
                {
                    "leads_sp": leads_sp_path,
                    "leads_pt": leads_pt_path,
                    "model_eq": model_equivalence_path,
                }
            )
            engine.OUTPUT_XLSX_PATH = output_path
            engine.main("01/08/2026", "01/08/2026")

            workbook = load_workbook(output_path, data_only=True)
            try:
                for sheet_name in ("LEADS_SP", "LEADS_PT"):
                    rows = list(workbook[sheet_name].iter_rows(values_only=True))
                    self.assertEqual(("model", "metric"), rows[0])
                    self.assertEqual(
                        ["A Model", "Z Model"],
                        [row[0] for row in rows if row[1] == "dealer leads"],
                    )
                    self.assertNotIn("OLD", [row[0] for row in rows])
            finally:
                workbook.close()

    def test_app_registers_the_leads_process_with_its_three_inputs(self):
        self.assertIn(app.PROCESS_LEADS_ANALYSIS, app.PROCESS_CONFIGS)
        self.assertEqual(
            ["leads_sp", "leads_pt", "model_eq"],
            app.PROCESS_CONFIGS[app.PROCESS_LEADS_ANALYSIS]["file_keys"],
        )

"""Canonical VehicleTracking reader shared by portfolio engines."""

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

import tabular_normalization as tabular
import vehicle_tracking_cache as tracking_cache


HEADER_ALIASES = {
    "vin": ["vin #", "vin"],
    "material_code": ["vehicle material code", "material code", "material_code"],
    "description": ["description(local)", "description"],
    "eta": ["purchase eta", "eta"],
    "port": ["current warehouse", "port"],
    "vessel": ["name of vessel", "vessel"],
    "dsn": ["delivery short name", "dsn"],
    "sap": ["customer#", "customer# (sap code)", "sap"],
    "gate_in": ["gate in date", "gate in", "gate_in"],
    "gate_out": ["gate out date", "actual gate out"],
    "production_date": ["production date"],
    "status": ["current status", "status"],
    "invoice_date": ["sales invoice date"],
    "customer_country": ["customer country"],
    "address": ["detail delivery address", "address"],
    "city": ["city"],
    "tag": ["tag name", "tag"],
    "related_order": ["related order#", "related order", "related_order"],
    "reserved_so": ["reserved so #", "reserved so", "reserved_so"],
    "dn_create_time": ["dn create time", "dn_create_time"],
    "allocation_date": ["allocation date", "allocate date", "allocation_date"],
}

RECORD_SOURCE_FIELDS = {
    "vin": "vin",
    "material_code": "material_code",
    "description": "description",
    "eta": "eta",
    "port": "port",
    "vessel_name": "vessel",
    "dsn": "dsn",
    "sap": "sap",
    "gate_in": "gate_in",
    "gate_out": "gate_out",
    "production_date": "production_date",
    "status": "status",
    "invoice_date": "invoice_date",
    "country": "customer_country",
    "address": "address",
    "city": "city",
    "tag": "tag",
    "related_order": "related_order",
    "reserved_so": "reserved_so",
    "dn_create_time": "dn_create_time",
    "allocation_date": "allocation_date",
}

ALL_SOURCE_FIELDS = tuple(dict.fromkeys(RECORD_SOURCE_FIELDS.values()))


def load_vehicle_tracking(
    source_path,
    cache_path,
    report_progress,
    required_fields=(),
    shipping_eta_overrides=None,
):
    """Load normalized records and validate required source headers on each run."""
    source_path = Path(source_path)
    signature = tracking_cache.vehicle_tracking_file_signature(source_path)
    cached = tracking_cache.load_cache(cache_path, signature, report_progress, required_fields)
    if cached is not None:
        return tracking_cache.apply_shipping_eta_overrides(
            cached,
            shipping_eta_overrides or {},
            report_progress,
        )

    report_progress("Reading VehicleTracking.xlsx...")
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        worksheet.reset_dimensions()
        try:
            headers = [tabular.format_value(value) for value in next(worksheet.iter_rows(values_only=True))]
        except StopIteration as exc:
            raise ValueError("VehicleTracking worksheet is empty.") from exc

        optional_fields = [field for field in ALL_SOURCE_FIELDS if field not in required_fields]
        indexes = tabular.build_indexes(
            headers,
            HEADER_ALIASES,
            required_fields,
            optional_fields,
        )
        max_column = tabular.max_required_col(indexes)
        report_progress(f"VehicleTracking.xlsx columns limited to {max_column} of {len(headers)}")

        by_vin = {}
        vin_counter = Counter()
        scanned_rows = 0
        for scanned_rows, row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=max_column, values_only=True),
            start=1,
        ):
            if scanned_rows % 25000 == 0:
                report_progress(
                    f"Reading VehicleTracking.xlsx: {scanned_rows:,} rows scanned, "
                    f"{len(by_vin):,} VINs loaded"
                )
            if not any(not tabular.is_missing(value) for value in row):
                continue

            vin = tabular.vin_key(tabular.row_value(row, indexes, "vin"))
            if not vin:
                continue

            vin_counter[vin] += 1
            by_vin[vin] = {
                "vin": vin,
                "material_code": tabular.code_key(tabular.row_value(row, indexes, "material_code")),
                "description": tabular.format_value(tabular.row_value(row, indexes, "description")),
                "eta": tabular.row_value(row, indexes, "eta"),
                "port": tabular.format_value(tabular.row_value(row, indexes, "port")),
                "vessel_name": tabular.format_value(tabular.row_value(row, indexes, "vessel")),
                "dsn": tabular.format_value(tabular.row_value(row, indexes, "dsn")),
                "sap": tabular.code_key(tabular.row_value(row, indexes, "sap")),
                "gate_in": tabular.row_value(row, indexes, "gate_in"),
                "gate_out": tabular.row_value(row, indexes, "gate_out"),
                "production_date": tabular.row_value(row, indexes, "production_date"),
                "status": tabular.format_value(tabular.row_value(row, indexes, "status")),
                "invoice_date": tabular.row_value(row, indexes, "invoice_date"),
                "country": tabular.format_value(tabular.row_value(row, indexes, "customer_country")),
                "address": tabular.format_value(tabular.row_value(row, indexes, "address")),
                "city": tabular.format_value(tabular.row_value(row, indexes, "city")),
                "tag": tabular.format_value(tabular.row_value(row, indexes, "tag")),
                "related_order": tabular.format_value(tabular.row_value(row, indexes, "related_order")),
                "reserved_so": tabular.format_value(tabular.row_value(row, indexes, "reserved_so")),
                "dn_create_time": tabular.row_value(row, indexes, "dn_create_time"),
                "allocation_date": tabular.row_value(row, indexes, "allocation_date"),
            }
    finally:
        workbook.close()

    duplicated = sorted(vin for vin, count in vin_counter.items() if count > 1)
    if duplicated:
        raise ValueError(f"vehicle_tracking.vin duplicated values: {', '.join(duplicated[:10])}")

    available_fields = tuple(field for field, index in indexes.items() if index is not None)
    report_progress(f"Vehicle tracking loaded: {len(by_vin):,} VINs from {scanned_rows:,} rows")
    tracking_cache.write_cache(cache_path, signature, by_vin, report_progress, available_fields)
    return tracking_cache.apply_shipping_eta_overrides(
        by_vin,
        shipping_eta_overrides or {},
        report_progress,
    )
